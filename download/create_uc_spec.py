"""
Generate user_creation_master_spec.xlsx — the authoritative spec for User Creation Screen automation.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.properties.creator = "Z.ai"

# ── Style helpers ──
hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sub_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = wrap
        cell.border = thin_border

def style_subheader(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = wrap
        cell.border = thin_border

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = wrap
        cell.border = thin_border

def auto_width(ws, cols, max_w=50):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = max_w

# ═══════════════════════════════════════════════════════
# SHEET 1: Fields
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Fields"
ws1.merge_cells("A1:N1")
ws1["A1"] = "User Creation Screen — Complete Field Inventory"
ws1["A1"].font = Font(bold=True, size=14, color="2F5496")

headers = ["#", "Field Label", "Field Type", "Required", "formControlName",
           "HTML Tag", "Default Value", "Placeholder", "maxlength",
           "Allowed Chars", "Dropdown Options", "Dependencies",
           "Read-only (View)", "Tab Order"]
r = 3
for ci, h in enumerate(headers, 1):
    ws1.cell(row=r, column=ci, value=h)
style_header(ws1, r, len(headers))

fields_data = [
    [1, "Username", "text input", "Yes *", "username", "input",
     "", "", "None (BUG: accepts 256+)",
     "No spaces allowed. Special chars trigger 'should not contain spaces'. Alphanums OK.",
     "N/A", "Must be unique (silently blocked if duplicate — BUG: no error msg)",
     "Yes (disabled)", "1"],
    [2, "Email", "text input", "Yes *", "email", "input",
     "", "", "None",
     "Accepts any text including invalid format (BUG: no email format validation on blur)",
     "N/A", "Can be reused across users (confirmed by user)",
     "Yes (disabled)", "2"],
    [3, "First Name", "text input", "Yes *", "first_name", "input",
     "", "", "None (BUG: no maxlength)",
     "Accepts special chars, no sanitization",
     "N/A", "None",
     "Yes (disabled)", "3"],
    [4, "Last Name", "text input", "Yes *", "last_name", "input",
     "", "", "None (BUG: no maxlength)",
     "Accepts special chars, no sanitization",
     "N/A", "None",
     "Yes (disabled)", "4"],
    [5, "Password", "password input", "Yes * (Create)", "password", "input[type=password]",
     "", "Leave blank to keep current", "None",
     "Any characters accepted",
     "N/A", "In Edit mode: empty = keep current. Placeholder changes to 'Leave blank to keep current'",
     "Yes (disabled)", "5"],
    [6, "User Type", "mat-select (searchable)", "Yes *", "—", "mat-select",
     "—", "—", "N/A",
     "Must select one option",
     "Maker, Checker, Both(Maker & Checker), Approver",
     "None observed",
     "Yes (disabled)", "6"],
    [7, "Role", "mat-select (searchable)", "Yes *", "—", "mat-select",
     "—", "—", "N/A",
     "Dynamic list from Role Creation Screen. 91 options at exploration.",
     "Dynamic (loaded from API). Includes: Agdi Admin, Rular Access, FPO/FPC Admin, etc.",
     "None observed",
     "Yes (disabled)", "7"],
    [8, "Entity", "mat-select (searchable)", "Yes *", "—", "mat-select",
     "—", "—", "N/A",
     "Dynamic list. 24 options at exploration.",
     "Dynamic (loaded from API). Includes: Agdi, Agri zone Farmer Producer Company, Height Green Producer Company, etc.",
     "None observed",
     "Yes (disabled)", "8"],
    [9, "Designation", "mat-select (searchable)", "Yes *", "—", "mat-select",
     "—", "—", "N/A",
     "4 options (BUG: 'Manager' appears twice)",
     "Manager, fr, AutoDesig DDXSBMAE, Manager",
     "None observed",
     "Yes (disabled)", "9"],
    [10, "Active", "mat-checkbox", "Optional", "is_active", "mat-checkbox",
     "Checked", "—", "N/A",
     "Toggle on/off",
     "N/A", "Default checked on Create form",
     "Yes (disabled)", "10"],
    [11, "Staff", "mat-checkbox", "Optional", "is_staff", "mat-checkbox",
     "Unchecked", "—", "N/A",
     "Toggle on/off",
     "N/A", "Default unchecked on Create form",
     "Yes (disabled)", "11"],
]

for i, row_data in enumerate(fields_data):
    r = 4 + i
    for ci, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=ci, value=val)
    style_data(ws1, r, len(headers))

auto_width(ws1, len(headers))

# ═══════════════════════════════════════════════════════
# SHEET 2: Validations
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Validations")
ws2.merge_cells("A1:G1")
ws2["A1"] = "User Creation Screen — Validation Matrix"
ws2["A1"].font = Font(bold=True, size=14, color="2F5496")

val_headers = ["#", "Validation", "Trigger", "Error Message", "Error Type", "Field Indicator", "Status"]
r = 3
# Sub-header: Required Field Validations
ws2.merge_cells(f"A{r}:G{r}")
ws2.cell(row=r, column=1, value="REQUIRED FIELD VALIDATIONS")
style_subheader(ws2, r, len(val_headers))

r = 4
for ci, h in enumerate(val_headers, 1):
    ws2.cell(row=r, column=ci, value=h)
style_header(ws2, r, len(val_headers))

validations = [
    [1, "Username required", "Submit/Update with empty Username", "Username is required", "mat-error (inline)", "input[formcontrolname=username] → ng-invalid, mat-form-field-invalid", "Working"],
    [2, "Email required", "Submit/Update with empty Email", "No visible error text (only CSS ng-invalid)", "CSS class only", "input[formcontrolname=email] → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [3, "First Name required", "Submit/Update with empty First Name", "No visible error text (only CSS ng-invalid)", "CSS class only", "input[formcontrolname=first_name] → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [4, "Last Name required", "Submit/Update with empty Last Name", "No visible error text (only CSS ng-invalid)", "CSS class only", "input[formcontrolname=last_name] → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [5, "Password required", "Submit/Update with empty Password", "No visible error text (only CSS ng-invalid)", "CSS class only", "input[formcontrolname=password] → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [6, "User Type required", "Submit/Update without selecting", "No visible error text (only CSS ng-invalid)", "CSS class only", "mat-select → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [7, "Role required", "Submit/Update without selecting", "No visible error text (only CSS ng-invalid)", "CSS class only", "mat-select → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [8, "Entity required", "Submit/Update without selecting", "No visible error text (only CSS ng-invalid)", "CSS class only", "mat-select → ng-invalid", "Partial — no inline text, only CSS highlight"],
    [9, "Designation required", "Submit/Update without selecting", "No visible error text (only CSS ng-invalid)", "CSS class only", "mat-select → ng-invalid", "Partial — no inline text, only CSS highlight"],
]

for i, row_data in enumerate(validations):
    r = 5 + i
    for ci, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=ci, value=val)
    style_data(ws2, r, len(val_headers))

# Sub-header: Additional Validations
r = 5 + len(validations) + 1
ws2.merge_cells(f"A{r}:G{r}")
ws2.cell(row=r, column=1, value="ADDITIONAL FIELD VALIDATIONS")
style_subheader(ws2, r, len(val_headers))

r += 1
for ci, h in enumerate(val_headers, 1):
    ws2.cell(row=r, column=ci, value=h)
style_header(ws2, r, len(val_headers))

additional_vals = [
    [10, "Username no spaces", "Type spaces in Username + blur", "Username should not contain spaces", "mat-error (inline)", "input[formcontrolname=username] → ng-invalid", "Working"],
    [11, "Username special chars", "Type !@#$%^&*() in Username + blur", "Username should not contain spaces (same message as spaces)", "mat-error (inline)", "input[formcontrolname=username] → ng-invalid", "BUG: Misleading error message for special chars"],
    [12, "Username unique (Create)", "Submit with existing username", "NO ERROR MESSAGE — form stays open silently", "SILENT BLOCK", "Form stays open, no SweetAlert2, no mat-error", "BUG: No feedback to user"],
    [13, "Username unique (Edit)", "Edit to existing username + Update", "TBD — likely same silent block", "SILENT BLOCK (suspected)", "TBD", "BUG: Needs verification"],
    [14, "Email format validation", "Type 'notanemail' + blur", "No error on blur (may validate on submit)", "None on blur", "No CSS change on blur", "BUG: No client-side email format validation"],
    [15, "First Name special chars", "Type !@#$%^&*() + blur", "No error", "None", "No validation", "BUG: No input sanitization"],
    [16, "Last Name special chars", "Type !@#$%^&*() + blur", "No error", "None", "No validation", "BUG: No input sanitization"],
    [17, "Username max length", "Type 256+ characters", "No error — all 256 chars accepted", "None", "No maxlength attribute", "BUG: No maxlength validation"],
    [18, "Active/Staff checkboxes optional", "Submit without toggling", "No error (correctly optional)", "N/A", "N/A", "Correctly optional"],
]

for i, row_data in enumerate(additional_vals):
    r2 = r + 1 + i
    for ci, val in enumerate(row_data, 1):
        ws2.cell(row=r2, column=ci, value=val)
    style_data(ws2, r2, len(val_headers))

auto_width(ws2, len(val_headers))

# ═══════════════════════════════════════════════════════
# SHEET 3: Bugs
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Bugs")
ws3.merge_cells("A1:H1")
ws3["A1"] = "User Creation Screen — Bug Registry"
ws3["A1"].font = Font(bold=True, size=14, color="2F5496")

bug_headers = ["Bug #", "Phase", "Description", "Expected Result", "Actual Result", "Severity", "Test #", "Status"]
r = 3
for ci, h in enumerate(bug_headers, 1):
    ws3.cell(row=r, column=ci, value=h)
style_header(ws3, r, len(bug_headers))

bugs = [
    [1, "Create", "Duplicate Username silently blocked — NO error message",
     "System should show validation error 'Username already exists'",
     "Form stays open on Submit with no SweetAlert2, no mat-error, no toast — completely silent",
     "High", "UC-D01", "Open"],
    [2, "Create", "No maxlength on Username — accepts 256+ characters",
     "System should restrict or truncate at a reasonable length (e.g., 255 chars)",
     "256+ character usernames accepted without any error or truncation",
     "Medium", "UC-C11", "Open"],
    [3, "Create/Edit", "No email format validation on blur or submit",
     "System should validate email format and show error for invalid emails",
     "Emails like 'notanemail' are accepted — no format validation anywhere",
     "High", "UC-C12", "Open"],
    [4, "Create", "Misleading error for special chars in Username",
     "Error message should say 'Username contains invalid characters'",
     "Error says 'Username should not contain spaces' even for !@#$%^&*()",
     "Low", "UC-C08", "Open"],
    [5, "Create/Edit", "No input sanitization on First Name / Last Name",
     "System should sanitize or reject special characters in name fields",
     "Special chars like !@#$%^&*() accepted without any validation",
     "Medium", "UC-C09, UC-C10", "Open"],
    [6, "Create/Edit", "Only 1 mat-error visible at a time",
     "All invalid fields should show their respective inline error messages",
     "Only Username shows mat-error text. Other fields get ng-invalid CSS but no visible text",
     "Medium", "UC-C01", "Open"],
    [7, "Create/Edit", "Duplicate 'Manager' option in Designation dropdown",
     "Dropdown should show unique options only",
     "Manager appears twice in the Designation dropdown options",
     "Low", "UC-P06", "Open"],
    [8, "Create", "No duplicate email prevention — email can be reused",
     "System may want to warn about duplicate emails",
     "Duplicate emails are freely accepted (user confirmed this is intentional)",
     "Info", "UC-D03", "By Design"],
]

for i, row_data in enumerate(bugs):
    r2 = 4 + i
    for ci, val in enumerate(row_data, 1):
        ws3.cell(row=r2, column=ci, value=val)
    style_data(ws3, r2, len(bug_headers))

auto_width(ws3, len(bug_headers))

# ═══════════════════════════════════════════════════════
# SHEET 4: Selectors
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Selectors")
ws4.merge_cells("A1:G1")
ws4["A1"] = "User Creation Screen — Complete Selector Map"
ws4["A1"].font = Font(bold=True, size=14, color="2F5496")

sel_headers = ["#", "Element", "CSS Selector", "XPath", "Tag", "Key Attributes", "Notes"]
r = 3

# Sub: Toolbar
ws4.merge_cells(f"A{r}:G{r}")
ws4.cell(row=r, column=1, value="TOOLBAR BUTTONS")
style_subheader(ws4, r, len(sel_headers))

r = 4
for ci, h in enumerate(sel_headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(sel_headers))

toolbar_sels = [
    [1, "Add Button", "button.erp-add-btn", "//button[contains(@class,'erp-add-btn')]", "BUTTON", "class='erp-add-btn'", "Text: 'add Add User Creation Screen'"],
    [2, "Search Toggle", "button[mattooltip='Search'], button.search-btn", "//button[@mattooltip='Search']", "BUTTON", "mattooltip='Search', aria-label='Search'", "Click first to enable readonly search input"],
    [3, "Search Input", "input.search-bar-input", "//input[contains(@class,'search-bar-input')]", "INPUT", "placeholder='Search anything...', readonly by default", "Must click Search Toggle first, then remove readonly via JS"],
    [4, "Refresh Button", "button[mattooltip='Refresh']", "//button[@mattooltip='Refresh']", "BUTTON", "mattooltip='Refresh'", "Icon: refresh (material-icons)"],
    [5, "Filter Button", "button.filter-btn, button[mattooltip='Filters']", "//button[@mattooltip='Filters']", "BUTTON", "mattooltip='Filters', class='filter-btn'", "Opens filter panel"],
    [6, "More Button", "button[mattooltip='More']", "//button[@mattooltip='More']", "BUTTON", "mattooltip='More'", "Opens dropdown menu (Export, Print, etc.)"],
]

for i, row_data in enumerate(toolbar_sels):
    r2 = 5 + i
    for ci, val in enumerate(row_data, 1):
        ws4.cell(row=r2, column=ci, value=val)
    style_data(ws4, r2, len(sel_headers))

# Sub: Table
r = 5 + len(toolbar_sels) + 1
ws4.merge_cells(f"A{r}:G{r}")
ws4.cell(row=r, column=1, value="TABLE")
style_subheader(ws4, r, len(sel_headers))

r += 1
for ci, h in enumerate(sel_headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(sel_headers))

table_sels = [
    [7, "Table", "table.mat-mdc-table", "//table[contains(@class,'mat-mdc-table')]", "TABLE", "class='mat-mdc-table mdc-data-table__table cdk-table mat-sort'", "No id attribute"],
    [8, "Table Rows", "table.mat-mdc-table tbody tr", "//table//tbody/tr", "TR", "", "Data rows only (exclude header)"],
    [9, "Username Cells", "td.cdk-column-username, td.mat-column-username", "//td[contains(@class,'cdk-column-username')]", "TD", "", "2nd column (after Actions)"],
    [10, "Email Cells", "td.cdk-column-email, td.mat-column-email", "//td[contains(@class,'cdk-column-email')]", "TD", "", "3rd column"],
    [11, "Joined Cells", "td.cdk-column-joined, td.mat-column-joined", "//td[contains(@class,'cdk-column-joined')]", "TD", "", "4th column — date format: 'DD Mon YYYY HH:MM AM/PM'"],
    [12, "Status Cells", "td.cdk-column-status, td.mat-column-status", "//td[contains(@class,'cdk-column-status')]", "TD", "", "5th column — 'Active' or 'Inactive'"],
    [13, "View Button (per row)", "app-feather-icons[icon='eye']", "//app-feather-icons[@icon='eye']/ancestor::button", "BUTTON", "icon='eye'", "1st action button in Actions column"],
    [14, "Edit Button (per row)", "app-feather-icons[icon='edit']", "//app-feather-icons[@icon='edit']/ancestor::button", "BUTTON", "icon='edit'", "2nd action button in Actions column"],
    [15, "History Button (per row)", "app-feather-icons[icon='clock']", "//app-feather-icons[@icon='clock']/ancestor::button", "BUTTON", "icon='clock'", "3rd action button in Actions column"],
    [16, "No Data Row", "tr.mat-mdc-no-data-row, td.no-data", "//tbody//tr[@class*='no-data' or td[@class*='no-data']]", "TR/TD", "", "Shown when table is empty"],
    [17, "Pagination", "mat-paginator", "//mat-paginator", "MAT-PAGINATOR", "", "Items per page, navigation buttons"],
]

for i, row_data in enumerate(table_sels):
    r2 = r + 1 + i
    for ci, val in enumerate(row_data, 1):
        ws4.cell(row=r2, column=ci, value=val)
    style_data(ws4, r2, len(sel_headers))

# Sub: Form Popup
r = r + 1 + len(table_sels) + 1
ws4.merge_cells(f"A{r}:G{r}")
ws4.cell(row=r, column=1, value="ADD / EDIT / VIEW POPUP")
style_subheader(ws4, r, len(sel_headers))

r += 1
for ci, h in enumerate(sel_headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(sel_headers))

popup_sels = [
    [18, "Popup Container", ".edit_pop_up.override_edit_pop_up", "//div[contains(@class,'edit_pop_up') and contains(@class,'override_edit_pop_up')]", "DIV", "class='edit_pop_up override_edit_pop_up'", "Contains .big-model child"],
    [19, "Popup Heading", "h3.popup-title", "//div[contains(@class,'edit_pop_up')]//h3[contains(@class,'popup-title')]", "H3", "", "Text: 'User Creation Screen Details'"],
    [20, "Close (X) Button", ".popup-actions button mat-icon[text()='close']", "//div[contains(@class,'popup-actions')]/button[mat-icon[text()='close']]", "BUTTON", "mat-icon: close", "Top-right corner of popup header"],
    [21, "Fullscreen Button", ".popup-actions button mat-icon[text()='fullscreen']", "//div[contains(@class,'popup-actions')]/button[mat-icon[text()='fullscreen']]", "BUTTON", "mat-icon: fullscreen", "Toggle popup fullscreen"],
    [22, "Username Input", "input[formcontrolname='username']", "//input[@formcontrolname='username']", "INPUT", "formcontrolname='username', required", "Text input, no spaces allowed"],
    [23, "Email Input", "input[formcontrolname='email']", "//input[@formcontrolname='email']", "INPUT", "formcontrolname='email', required", "Text input, no format validation on blur"],
    [24, "First Name Input", "input[formcontrolname='first_name']", "//input[@formcontrolname='first_name']", "INPUT", "formcontrolname='first_name', required", "Text input"],
    [25, "Last Name Input", "input[formcontrolname='last_name']", "//input[@formcontrolname='last_name']", "INPUT", "formcontrolname='last_name', required", "Text input"],
    [26, "Password Input", "input[formcontrolname='password']", "//input[@formcontrolname='password']", "INPUT", "formcontrolname='password', type='password', required", "Placeholder: 'Leave blank to keep current' in Edit"],
    [27, "User Type Select", "mat-label[text()='User Type']/ancestor::mat-form-field//mat-select", "//mat-label[contains(.,'User Type')]/ancestor::mat-form-field//mat-select", "MAT-SELECT", "required, searchable", "Options: Maker, Checker, Both(Maker & Checker), Approver"],
    [28, "Role Select", "mat-label[text()='Role']/ancestor::mat-form-field//mat-select", "//mat-label[contains(.,'Role')]/ancestor::mat-form-field//mat-select", "MAT-SELECT", "required, searchable", "Dynamic options from API"],
    [29, "Entity Select", "mat-label[text()='Entity']/ancestor::mat-form-field//mat-select", "//mat-label[contains(.,'Entity')]/ancestor::mat-form-field//mat-select", "MAT-SELECT", "required, searchable", "Dynamic options from API"],
    [30, "Designation Select", "mat-label[text()='Designation']/ancestor::mat-form-field//mat-select", "//mat-label[contains(.,'Designation')]/ancestor::mat-form-field//mat-select", "MAT-SELECT", "required, searchable", "Options: Manager, fr, AutoDesig DDXSBMAE, Manager"],
    [31, "Active Checkbox", "mat-checkbox[formcontrolname='is_active']", "//mat-checkbox[@formcontrolname='is_active']", "MAT-CHECKBOX", "formcontrolname='is_active'", "Default: checked"],
    [32, "Staff Checkbox", "mat-checkbox[formcontrolname='is_staff']", "//mat-checkbox[@formcontrolname='is_staff']", "MAT-CHECKBOX", "formcontrolname='is_staff'", "Default: unchecked"],
    [33, "Submit Button", ".popup-footer button.mdc-button--raised", "//div[@class='popup-footer']//button[contains(.,'Submit')]", "BUTTON", "class='mdc-button--raised mat-primary'", "Create mode only"],
    [34, "Update Button", ".popup-footer button.mdc-button--raised", "//div[@class='popup-footer']//button[contains(.,'Update')]", "BUTTON", "class='mdc-button--raised mat-primary'", "Edit mode only"],
    [35, "Cancel Button", ".popup-footer button:first-child", "//div[@class='popup-footer']//button[contains(.,'Cancel')]", "BUTTON", "", "Present in Add/Edit/View modes"],
]

for i, row_data in enumerate(popup_sels):
    r2 = r + 1 + i
    for ci, val in enumerate(row_data, 1):
        ws4.cell(row=r2, column=ci, value=val)
    style_data(ws4, r2, len(sel_headers))

# Sub: SweetAlert2 & History
r = r + 1 + len(popup_sels) + 1
ws4.merge_cells(f"A{r}:G{r}")
ws4.cell(row=r, column=1, value="SWEETALERT2 & HISTORY POPUP")
style_subheader(ws4, r, len(sel_headers))

r += 1
for ci, h in enumerate(sel_headers, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header(ws4, r, len(sel_headers))

other_sels = [
    [36, "SWAL2 Title", "#swal2-title", "//*[@id='swal2-title']", "H2", "id='swal2-title'", "Success: 'User Created Successfully' (expected)"],
    [37, "SWAL2 Content", ".swal2-content, .swal2-html-container", "//*[contains(@class,'swal2-content')]", "DIV", "", "May contain detail text"],
    [38, "SWAL2 Confirm", ".swal2-confirm", "//*[contains(@class,'swal2-confirm')]", "BUTTON", "class='swal2-confirm'", "OK button on success/warning popup"],
    [39, "SWAL2 Cancel", ".swal2-cancel", "//*[contains(@class,'swal2-cancel')]", "BUTTON", "class='swal2-cancel'", "Cancel button on warning popup"],
    [40, "History Popup", "h3.popup-title[text()='User Creation Screen History']", "//div[contains(@class,'big-model')][.//h3[contains(.,'History')]]", "DIV", "", "History popup container"],
    [41, "History Search Input", ".edit_pop_up input[placeholder='Search in table']", "//div[contains(@class,'edit_pop_up')]//input[@placeholder='Search in table']", "INPUT", "placeholder='Search in table'", "Search within history table"],
    [42, "History Close Button", ".popup-footer button[contains(.,'Close')]", "//div[@class='popup-footer']//button[contains(.,'Close')]", "BUTTON", "", "Closes history popup"],
    [43, "Validation Error", "mat-error, .mat-mdc-form-field-error", "//mat-error", "MAT-ERROR", "", "Inline field validation errors"],
    [44, "Dropdown Panel", "div.cdk-overlay-pane mat-select-panel", "//div[contains(@class,'cdk-overlay-pane')]//mat-select-panel", "DIV", "role='listbox'", "Angular Material dropdown overlay"],
    [45, "Dropdown Options", "div[role='listbox'] mat-option", "//div[@role='listbox']//mat-option", "MAT-OPTION", "role='option'", "Individual dropdown options"],
]

for i, row_data in enumerate(other_sels):
    r2 = r + 1 + i
    for ci, val in enumerate(row_data, 1):
        ws4.cell(row=r2, column=ci, value=val)
    style_data(ws4, r2, len(sel_headers))

auto_width(ws4, len(sel_headers))

# ═══════════════════════════════════════════════════════
# SHEET 5: Workflows
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Workflows")
ws5.merge_cells("A1:D1")
ws5["A1"] = "User Creation Screen — Step-by-Step Action Flows"
ws5["A1"].font = Font(bold=True, size=14, color="2F5496")

wf_headers = ["Step", "Action", "Selector / Code", "Notes"]
r = 3

# Create workflow
ws5.merge_cells(f"A{r}:D{r}")
ws5.cell(row=r, column=1, value="CREATE (Add New User)")
style_subheader(ws5, r, len(wf_headers))
r += 1
for ci, h in enumerate(wf_headers, 1):
    ws5.cell(row=r, column=ci, value=h)
style_header(ws5, r, len(wf_headers))

create_steps = [
    [1, "Navigate to page", "page.navigate_to_page()", "Calls driver.refresh() to clear SPA state"],
    [2, "Click ADD button", "page.open_add_form()", "button.erp-add-btn → popup opens with heading 'User Creation Screen Details'"],
    [3, "Fill Username", "page.type_text(USERNAME_INPUT, data['username'])", "input[formcontrolname='username'] — no spaces allowed"],
    [4, "Fill Email", "page.type_text(EMAIL_INPUT, data['email'])", "input[formcontrolname='email'] — no format validation on blur"],
    [5, "Fill First Name", "page.type_text(FIRST_NAME_INPUT, data['first_name'])", "input[formcontrolname='first_name']"],
    [6, "Fill Last Name", "page.type_text(LAST_NAME_INPUT, data['last_name'])", "input[formcontrolname='last_name']"],
    [7, "Fill Password", "page.type_text(PASSWORD_INPUT, data['password'])", "input[formcontrolname='password'], type=password"],
    [8, "Select User Type", "page._select_mat_option(USER_TYPE_SELECT, 'Maker')", "mat-select for 'User Type' → click → select option"],
    [9, "Select Role", "page._select_mat_option(ROLE_SELECT, 'Agdi Admin')", "Dynamic dropdown — use _select_random_from_dropdown if no specific role needed"],
    [10, "Select Entity", "page._select_mat_option(ENTITY_SELECT, 'Agdi')", "Dynamic dropdown — use _select_random_from_dropdown if no specific entity needed"],
    [11, "Select Designation", "page._select_mat_option(DESIGNATION_SELECT, 'Manager')", "4 options, use _select_random_from_dropdown"],
    [12, "Close dropdown panels", "page._force_close_panels()", "Remove all cdk-overlay panes (not dialog backdrop)"],
    [13, "Click Submit", "page.submit()", "button.mdc-button--raised containing 'Submit'"],
    [14, "Handle success alert", "page.handle_success_alert()", "SweetAlert2 '#swal2-title' → click .swal2-confirm"],
    [15, "Verify in table", "page.is_user_in_table(username)", "Refresh table, search for username"],
]

for i, row_data in enumerate(create_steps):
    r2 = r + 1 + i
    for ci, val in enumerate(row_data, 1):
        ws5.cell(row=r2, column=ci, value=val)
    style_data(ws5, r2, len(wf_headers))

# Edit workflow
r = r + 1 + len(create_steps) + 1
ws5.merge_cells(f"A{r}:D{r}")
ws5.cell(row=r, column=1, value="EDIT (Update Existing User)")
style_subheader(ws5, r, len(wf_headers))
r += 1
for ci, h in enumerate(wf_headers, 1):
    ws5.cell(row=r, column=ci, value=h)
style_header(ws5, r, len(wf_headers))

edit_steps = [
    [1, "Click Edit on row", "page.click_edit_button(username)", "app-feather-icons[icon='edit'] → ancestor::button"],
    [2, "Verify form pre-populated", "page.get_form_field_values()", "All fields should contain existing data; Password empty with placeholder"],
    [3, "Modify fields", "page.type_text(field, new_value, clear_first=True)", "Only change needed fields"],
    [4, "Close dropdown panels", "page._force_close_panels()", "Remove overlay panes"],
    [5, "Click Update", "page.click_update()", "button.mdc-button--raised containing 'Update'"],
    [6, "Handle success alert", "page.handle_success_alert()", "SweetAlert2 confirm"],
]

for i, row_data in enumerate(edit_steps):
    r2 = r + 1 + i
    for ci, val in enumerate(row_data, 1):
        ws5.cell(row=r2, column=ci, value=val)
    style_data(ws5, r2, len(wf_headers))

# View, Search, History workflows
for wf_name, wf_steps in [
    ("VIEW (Read-Only Popup)", [
        [1, "Click View on row", "page.click_view_button(username)", "app-feather-icons[icon='eye'] → ancestor::button"],
        [2, "Verify all fields disabled", "page.verify_view_popup_read_only()", "All inputs disabled, no Submit/Update button, only Cancel"],
        [3, "Close popup", "page.close_popup() or page.cancel()", "X button or Cancel button"],
    ]),
    ("SEARCH", [
        [1, "Click Search toggle", "page.click_search_toggle()", "button[mattooltip='Search'] → enables readonly input"],
        [2, "Type search term", "page.search_item(term)", "Remove readonly via JS, set value, press Enter"],
        [3, "Verify results", "page.get_all_usernames()", "Table should filter to matching rows"],
        [4, "Clear search", "page.clear_search()", "Clear input, press Enter to restore all rows"],
    ]),
    ("HISTORY", [
        [1, "Click History on row", "page.click_history_button(username)", "app-feather-icons[icon='clock'] → ancestor::button"],
        [2, "Verify history popup", "page.is_history_popup_open()", "Heading: 'User Creation Screen History'"],
        [3, "Search in history", "page.search_history(term)", "input[placeholder='Search in table'], press Enter"],
        [4, "Close history", "page.close_history_popup()", "Button containing 'Close'"],
    ]),
]:
    r = r + 1 + len(wf_steps) + 1
    ws5.merge_cells(f"A{r}:D{r}")
    ws5.cell(row=r, column=1, value=wf_name)
    style_subheader(ws5, r, len(wf_headers))
    r += 1
    for ci, h in enumerate(wf_headers, 1):
        ws5.cell(row=r, column=ci, value=h)
    style_header(ws5, r, len(wf_headers))
    for i, row_data in enumerate(wf_steps):
        r2 = r + 1 + i
        for ci, val in enumerate(row_data, 1):
            ws5.cell(row=r2, column=ci, value=val)
        style_data(ws5, r2, len(wf_headers))

auto_width(ws5, len(wf_headers))

# ═══════════════════════════════════════════════════════
# SHEET 6: Test Plan
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Test Plan")
ws6.merge_cells("A1:H1")
ws6["A1"] = "User Creation Screen — Proposed Test Plan"
ws6["A1"].font = Font(bold=True, size=14, color="2F5496")

tp_headers = ["#", "Test ID", "Phase", "Description", "Expected Result", "Bug #", "Class", "xfail"]
r = 3
for ci, h in enumerate(tp_headers, 1):
    ws6.cell(row=r, column=ci, value=h)
style_header(ws6, r, len(tp_headers))

test_plan = [
    # Phase C: Create
    [1, "UC-C01", "Create", "Submit with all fields empty", "Validation errors on all 8 required fields", "Bug #6", "TestCreateFormValidations", "No"],
    [2, "UC-C02", "Create", "Valid user creation (happy path)", "User created, success alert, appears in table", "", "TestCreateFormValidations", "No"],
    [3, "UC-C03", "Create", "Only Username filled — submit", "Validation errors on 7 remaining required fields", "Bug #6", "TestCreateFormValidations", "No"],
    [4, "UC-C04", "Create", "Username with spaces only", "Error: 'Username should not contain spaces'", "", "TestCreateFormValidations", "No"],
    [5, "UC-C05", "Create", "Username with leading/trailing spaces", "Spaces trimmed or rejected", "", "TestCreateFormValidations", "No"],
    [6, "UC-C06", "Create", "Email with invalid format", "Error: invalid email format", "Bug #3", "TestCreateFormValidations", "Yes"],
    [7, "UC-C07", "Create", "First Name with special characters", "Rejected or sanitized", "Bug #5", "TestCreateFormValidations", "Yes"],
    [8, "UC-C08", "Create", "Username with special characters", "Error about invalid characters", "Bug #4", "TestCreateFormValidations", "No"],
    [9, "UC-C09", "Create", "First Name with spaces only", "Rejected or trimmed", "", "TestCreateFormValidations", "No"],
    [10, "UC-C10", "Create", "Last Name with spaces only", "Rejected or trimmed", "", "TestCreateFormValidations", "No"],
    [11, "UC-C11", "Create", "Very long Username (256+ chars)", "Rejected or truncated", "Bug #2", "TestCreateFormValidations", "Yes"],
    [12, "UC-C12", "Create", "Email format validation on submit", "Invalid email rejected on submit", "Bug #3", "TestCreateFormValidations", "Yes"],
    [13, "UC-C13", "Create", "Without User Type selected", "Validation error on User Type", "Bug #6", "TestCreateFormValidations", "No"],
    [14, "UC-C14", "Create", "Without Role selected", "Validation error on Role", "Bug #6", "TestCreateFormValidations", "No"],
    [15, "UC-C15", "Create", "Without Entity selected", "Validation error on Entity", "Bug #6", "TestCreateFormValidations", "No"],
    [16, "UC-C16", "Create", "Without Designation selected", "Validation error on Designation", "Bug #6", "TestCreateFormValidations", "No"],
    # Phase D: Duplicate
    [17, "UC-D01", "Duplicate", "Duplicate username in Create", "Error: 'Username already exists'", "Bug #1", "TestDuplicateValidations", "Yes"],
    [18, "UC-D02", "Duplicate", "Duplicate email in Create", "Email accepted (by design)", "", "TestDuplicateValidations", "No"],
    [19, "UC-D03", "Duplicate", "Case-insensitive duplicate username", "Reject same name different case", "", "TestDuplicateValidations", "No"],
    # Phase E: Edit
    [20, "UC-E01", "Edit", "Edit with valid data (happy path)", "User updated, success alert, changes in table", "", "TestEditFormValidations", "No"],
    [21, "UC-E02", "Edit", "Edit pre-populated fields verification", "All fields pre-filled with existing data", "", "TestEditFormValidations", "No"],
    [23, "UC-E03", "Edit", "Edit with empty Username", "Validation error on Username", "", "TestEditFormValidations", "No"],
    [24, "UC-E04", "Edit", "Edit to duplicate Username", "Error: 'Username already exists'", "Bug #1", "TestEditFormValidations", "Yes"],
    [25, "UC-E05", "Edit", "Edit with invalid email format", "Error: invalid email", "Bug #3", "TestEditFormValidations", "Yes"],
    [26, "UC-E06", "Edit", "Edit password field (leave blank)", "Password unchanged when left blank", "", "TestEditFormValidations", "No"],
    # Phase S: Search
    [27, "UC-S01", "Search", "Search with exact username match", "User found in table", "", "TestSearchFilter", "No"],
    [28, "UC-S02", "Search", "Search with partial username", "Matching users shown", "", "TestSearchFilter", "No"],
    [29, "UC-S03", "Search", "Search for non-existent username", "No results / 'No data available'", "", "TestSearchFilter", "No"],
    [30, "UC-S04", "Search", "Search with special characters", "No crash, handled gracefully", "", "TestSearchFilter", "No"],
    # Phase P: Popup & UI
    [31, "UC-P01", "Popup", "Cancel discards form data", "Popup closes, no record created", "", "TestPopupUIBehaviors", "No"],
    [32, "UC-P02", "Popup", "X button closes form", "Popup closes, no record created", "", "TestPopupUIBehaviors", "No"],
    [33, "UC-P03", "Popup", "View popup shows read-only fields", "All fields disabled, only Cancel button", "", "TestPopupUIBehaviors", "No"],
    [34, "UC-P04", "Popup", "Edit popup shows Update button", "Update button visible, fields editable", "", "TestPopupUIBehaviors", "No"],
    [35, "UC-P05", "Popup", "Fullscreen button toggles popup", "Popup expands/collapses", "", "TestPopupUIBehaviors", "No"],
    [36, "UC-P06", "Popup", "Designation dropdown duplicate Manager", "Manager should appear only once", "Bug #7", "TestPopupUIBehaviors", "Yes"],
    # Phase H: History
    [37, "UC-H01", "History", "History popup opens after creation", "At least 1 history row", "", "TestHistoryValidations", "No"],
    [38, "UC-H02", "History", "History search with Enter key", "Search filters history rows", "", "TestHistoryValidations", "No"],
    [39, "UC-H03", "History", "History Close button works", "Popup closes", "", "TestHistoryValidations", "No"],
    # Phase B: Bug-specific
    [40, "UC-B01", "Bug", "Silent duplicate username block — no error message", "Error message displayed to user", "Bug #1", "TestBugSpecific", "Yes"],
    [41, "UC-B02", "Bug", "Only 1 mat-error visible at a time", "All invalid fields show inline errors", "Bug #6", "TestBugSpecific", "Yes"],
    [42, "UC-B03", "Bug", "No email format validation anywhere", "Email format validated on blur or submit", "Bug #3", "TestBugSpecific", "Yes"],
    [43, "UC-B04", "Bug", "No maxlength on text inputs", "Max length enforced or truncated", "Bug #2", "TestBugSpecific", "Yes"],
]

for i, row_data in enumerate(test_plan):
    r2 = 4 + i
    for ci, val in enumerate(row_data, 1):
        ws6.cell(row=r2, column=ci, value=val)
    style_data(ws6, r2, len(tp_headers))

auto_width(ws6, len(tp_headers))

# ── Save ──
output_path = "/home/z/my-project/download/user_creation_master_spec.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
