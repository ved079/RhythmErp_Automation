"""Batch data creation via the ERP API — SSE-streamed progress."""

import importlib
import inspect
import json
import logging
import subprocess
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Generator

from api.models import BatchCreateRequest, LogEvent

logger = logging.getLogger(__name__)

RESULTS_DIR = Path(__file__).parent / "batch_results"

MODULE_IMPORT_PATHS = {
    "registration": {
        "supplier": "pages.registration.modules.supplier.data.supplier_data",
        "farmer": "pages.registration.modules.farmer.data.farmer_data",
        "customer": "pages.registration.modules.customer.data.customer_data",
        "employee": "pages.registration.modules.employee.data.employee_data",
        "agent": "pages.registration.modules.agent.data.agent_data",
        "directors": "pages.documents.modules.directors.data.directors_data",
        "member": "pages.documents.modules.member.data.member_data",
        "miscellaneous_documents": "pages.documents.modules.miscellaneous_documents.data.miscellaneous_documents_data",
        "constituent_documents": "pages.documents.modules.constituent_documents.data.constituent_documents_data",
        "register_of_loan": "pages.documents.modules.register_of_loan.data.register_of_loan_data",
        "register_charges": "pages.documents.modules.register_charges.data.register_charges_data",
    },
    "common_settings": {
        "bank": "pages.common_settings.modules.bank.data.bank_data",
        "designation": "pages.common_settings.modules.designation.data.designation_data",
        "error_code_mst": "pages.common_settings.modules.error_code_mst.data.error_code_mst_data",
        "hsn_sac": "pages.common_settings.modules.hsn_sac.data.hsn_sac_data",
        "season": "pages.common_settings.modules.season.data.season_data",
        "tax_authority": "pages.common_settings.modules.tax_authority.data.tax_authority_data",
        "tax_rate": "pages.common_settings.modules.tax_rate.data.tax_rate_data",
        "uom": "pages.common_settings.modules.uom.data.uom_data",
        "uom_conversion": "pages.common_settings.modules.uom_conversion.data.uom_conversion_data",
        "vehicle_master": "pages.common_settings.modules.vehicle_master.data.vehicle_master_data",
    },
    "company_onboarding": {
        "company_onboarding": "pages.company_onboarding.data.company_onboarding_data",
    },
    "commodity_settings": {
        "item_category": "pages.commodity_settings.modules.item_category.data.item_category_data",
        "item_group": "pages.commodity_settings.modules.item_group.data.item_group_data",
        "item_master": "pages.commodity_settings.modules.item_master.data.item_master_data",
        "item_attribute": "pages.commodity_settings.modules.item_attribute.data.item_attribute_data",
        "crop_master": "pages.commodity_settings.modules.crop_master.data.crop_master_data",
        "services_master": "pages.commodity_settings.modules.services_master.data.services_master_data",
        "quality_parameter_master": "pages.commodity_settings.modules.quality_parameter_master.data.quality_parameter_master_data",
        "commodity_quality_parameter": "pages.commodity_settings.modules.commodity_quality_parameter.data.commodity_quality_parameter_data",
        "commodity_base_rate": "pages.commodity_settings.modules.commodity_base_rate.data.cbr_data",
    },
}


# Modules that use a standalone CLI script instead of generate_batch_payloads.
# Value is the path to the script relative to PROJECT_ROOT.
_SCRIPT_PATHS: dict[str, dict[str, Path]] = {
    "private_b2b": {
        "direct_pb_flow": Path("pages/private_b2b/modules/Purchase_Flow_Tests/test/playwright/direct_pb_flow/scripts/batch_create.py"),
        "po_qc_pb_flow": Path("pages/private_b2b/modules/Purchase_Flow_Tests/test/playwright/po_qc_pb_flow/scripts/batch_create.py"),
    },
}

PROJECT_ROOT = Path(__file__).parent.parent


def _sse_event(event: LogEvent) -> str:
    return f"data: {event.model_dump_json()}\n\n"


# Map sub_module keys → ERP screen attribute_name
_SCREEN_NAME_MAP: dict[str, str] = {
    "farmer": "Farmer",
    "supplier": "Supplier",
    "customer": "Customer",
    "employee": "Employee",
    "agent": "Agent",
    "directors": "Directors",
    "member": "Member",
    "miscellaneous_documents": "Miscellaneous Documents",
    "constituent_documents": "Constituent Documents",
    "register_of_loan": "Register of Loan",
    "register_charges": "Register Charges",
    "bank": "Bank",
    "designation": "Designation",
    "error_code_mst": "Error Code Mst",
    "hsn_sac": "HSN SAC",
    "season": "Season",
    "tax_authority": "Tax Authority",
    "tax_rate": "Tax Rate",
    "vehicle_master": "Vehicle Master",
    "uom": "UOM",
    "uom_conversion": "UOM Conversion",
    "item_category": "Item Category",
    "item_group": "Item Group",
    "item_master": "Item Master",
    "item_attribute": "Item Attribute",
    "crop_master": "Crop Master",
    "services_master": "Services Master",
    "quality_parameter_master": "Quality Parameter Master",
    "commodity_quality_parameter": "Commodity Quality Parameter",
    "commodity_base_rate": "Commodity Base Rate",
    "company_onboarding": "Company Onboarding",
}


def _screen_name(sub_module: str) -> str:
    return _SCREEN_NAME_MAP.get(sub_module.lower(), sub_module.capitalize())


def _compute_land_classification(payload: dict) -> int | None:
    """Return the land_classification ID based on individual_land_holding in the payload.

    Mirrors the ERP UI's JS calculation on Edit form load:
      644 = Marginal Farmer  (individual holding < 1 ha)
      645 = Small Farmer     (1 <= holding <= 2 ha)
      646 = Other Farmer     (holding > 2 ha)

    Returns None if no land details are present (e.g. walk-in).
    """
    for child in payload.get("children", []):
        if child.get("stepper_name") == "Land Details":
            rows = child.get("details", [])
            if rows:
                holding = rows[0].get("individual_land_holding") or 0
                try:
                    holding = float(holding)
                except (TypeError, ValueError):
                    holding = 0
                if holding < 1:
                    return 644  # Marginal
                elif holding <= 2:
                    return 645  # Small
                else:
                    return 646  # Other
    return None


def _extract_wf_error(client) -> str:
    last = getattr(client, "_last_raw_response", None)
    if last is None:
        return "unknown"
    try:
        body = last.json()
        return body.get("error") or body.get("message") or last.text[:200]
    except Exception:
        return last.text[:200]


def _build_payloads_only(request: BatchCreateRequest) -> list[dict]:
    """Generate payloads without creating them — used by conflict mode preview."""
    from common.erp_api_client import RhythmERPAPIClient

    module_paths = MODULE_IMPORT_PATHS.get(request.module, {})
    data_path = module_paths.get(request.sub_module)
    if not data_path:
        return []

    data_mod = importlib.import_module(data_path)
    generate_fn = getattr(data_mod, "generate_batch_payloads", None)
    if not generate_fn:
        return []

    client = RhythmERPAPIClient(username="", password="", tenant_id=request.erp_tenant_id)
    try:
        client.login_from_browser(token=request.erp_token, tenant_id=request.erp_tenant_id)
    except Exception:
        return []

    # FK resolution
    raw_fk_map = getattr(data_mod, "get_fk_screen_mapping", lambda: [])()
    if isinstance(raw_fk_map, list):
        fk_map = {item["fk_field"]: item["screen_name"] for item in raw_fk_map}
    elif isinstance(raw_fk_map, dict):
        fk_map = raw_fk_map
    else:
        fk_map = {}
    dropdown_ids = None
    if fk_map:
        from common.fk_resolver import FkResolver
        try:
            resolver = FkResolver(client)
            parent_screen = _screen_name(request.sub_module)
            dropdown_ids = {}
            for field, screen in fk_map.items():
                resolved = resolver.resolve(screen, parent_screen=parent_screen, field_key=field)
                if resolved:
                    dropdown_ids[field] = resolved
            if "item_ref_id" in dropdown_ids:
                try:
                    im_resp = client.list_entries("Item Master", page_size=500)
                    im_rows = im_resp.get("screenmatlistingdata_set", im_resp.get("results", im_resp.get("data", [])))
                    item_uom_map = {}
                    for row in (im_rows or []):
                        iid = row.get("id")
                        uom = row.get("uom") or row.get("base_uom")
                        if iid is not None and uom is not None:
                            item_uom_map[int(iid)] = uom
                    if item_uom_map:
                        dropdown_ids["item_uom_map"] = item_uom_map
                except Exception:
                    pass
        except Exception:
            pass

    # Query existing entries
    try:
        screen = _screen_name(request.sub_module)
        existing = client.list_entries(screen, page_size=500)
        existing_items = existing.get("screenmatlistingdata_set", existing.get("results", existing.get("data", [])))
    except Exception:
        existing_items = []

    enrich_fn = getattr(data_mod, "enrich_existing_entries", None)
    if enrich_fn and existing_items:
        try:
            existing_items = enrich_fn(client, existing_items)
        except Exception:
            pass

    sig = inspect.signature(generate_fn)
    total = min(request.count, 500)
    kwargs = {"count": total}
    if dropdown_ids and "dropdown_ids" in sig.parameters:
        kwargs["dropdown_ids"] = dropdown_ids
    if existing_items and "existing_entries" in sig.parameters:
        kwargs["existing_entries"] = existing_items
    if "config" in sig.parameters:
        kwargs["config"] = request.config
    attr_number = (request.config or {}).get("attr_number", 1)
    if attr_number and "attr_number" in sig.parameters:
        kwargs["attr_number"] = attr_number

    payloads = generate_fn(**kwargs)

    # Tenant-universal FK pins (e.g. Customer Sale Type = "Commission", resolved live)
    pin_fn = getattr(data_mod, "apply_tenant_fk_pins", None)
    if callable(pin_fn):
        try:
            payloads = pin_fn(client, payloads)
        except Exception:
            pass

    try:
        client.close()
    except Exception:
        pass

    return payloads


def _run_script_stream(script_path: Path, request: BatchCreateRequest, run_id: str) -> Generator[str, None, None]:
    """Run a standalone batch-create script as a subprocess and yield its stdout as SSE log events."""
    args = [
        sys.executable, str(PROJECT_ROOT / script_path),
        "--token",  request.erp_token,
        "--tenant", request.erp_tenant_id or "Eco Green Pvt Ltd",
        "--count",  str(request.count),
    ]
    config = request.config or {}
    if config.get("item_range"):
        args += ["--item-range", config["item_range"]]
    if config.get("items_per_pb"):
        args += ["--items-per-pb", str(config["items_per_pb"])]
    if config.get("items_per_chain"):
        args += ["--items-per-chain", str(config["items_per_chain"])]
    if config.get("supplier"):
        args += ["--supplier", str(config["supplier"])]

    yield _sse_event(LogEvent(
        type="log", run_id=run_id,
        message=f"Starting script: {script_path.name}  (count={request.count})",
        timestamp=datetime.now(timezone.utc),
    ))

    try:
        proc = subprocess.Popen(args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                text=True, cwd=str(PROJECT_ROOT))
        n_created = 0
        n_failed = 0
        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue
            if line.startswith("  [") and "CREATED #" in line:
                n_created += 1
                yield _sse_event(LogEvent(type="created", run_id=run_id, message=line.strip(),
                                          timestamp=datetime.now(timezone.utc)))
            elif line.startswith("  [") and ("FAILED" in line or "ERROR" in line):
                n_failed += 1
                yield _sse_event(LogEvent(type="failed", run_id=run_id, message=line.strip(),
                                          timestamp=datetime.now(timezone.utc)))
            else:
                yield _sse_event(LogEvent(type="log", run_id=run_id, message=line,
                                          timestamp=datetime.now(timezone.utc)))
        proc.wait()
        yield _sse_event(LogEvent(
            type="run_end", run_id=run_id,
            message=f"{n_created} created, {n_failed} failed",
            timestamp=datetime.now(timezone.utc),
        ))
    except Exception as e:
        yield _sse_event(LogEvent(type="error", run_id=run_id, message=str(e),
                                  timestamp=datetime.now(timezone.utc)))


def batch_create_stream(request: BatchCreateRequest) -> Generator[str, None, None]:
    """Generate batch create progress as SSE events."""
    run_id = str(uuid.uuid4())[:8]

    # Script-backed modules (standalone scripts with their own FK resolution)
    script_path = _SCRIPT_PATHS.get(request.module, {}).get(request.sub_module)
    if script_path:
        yield from _run_script_stream(script_path, request, run_id)
        return

    module_paths = MODULE_IMPORT_PATHS.get(request.module, {})
    data_path = module_paths.get(request.sub_module)
    if not data_path:
        yield _sse_event(LogEvent(
            type="error",
            message=f"Unsupported module: {request.module}/{request.sub_module}",
            timestamp=datetime.now(timezone.utc),
            run_id=run_id,
        ))
        return

    # 1. Dynamic import of data generators
    try:
        data_mod = importlib.import_module(data_path)
        generate_fn = getattr(data_mod, "generate_batch_payloads", None)
        if not generate_fn:
            yield _sse_event(LogEvent(
                type="error",
                message=f"No generate_batch_payloads found in {data_path}",
                timestamp=datetime.now(timezone.utc),
                run_id=run_id,
            ))
            return
    except Exception as e:
        yield _sse_event(LogEvent(
            type="error",
            message=f"Failed to import data module: {e}",
            timestamp=datetime.now(timezone.utc),
            run_id=run_id,
        ))
        return

    # 2. Create ERP API client
    from common.erp_api_client import RhythmERPAPIClient
    client = RhythmERPAPIClient(
        username="",
        password="",
        tenant_id=request.erp_tenant_id,
    )
    try:
        client.login_from_browser(token=request.erp_token, tenant_id=request.erp_tenant_id)
    except Exception as e:
        yield _sse_event(LogEvent(
            type="error",
            message=f"ERP API auth failed: {e}",
            timestamp=datetime.now(timezone.utc),
            run_id=run_id,
        ))
        return

    # 3. Resolve FK IDs if the module exposes a mapping
    raw_fk_map = getattr(data_mod, "get_fk_screen_mapping", lambda: [])()
    # Normalise: list of dicts → {fk_field: screen_name}
    if isinstance(raw_fk_map, list):
        fk_map = {item["fk_field"]: item["screen_name"] for item in raw_fk_map}
    elif isinstance(raw_fk_map, dict):
        fk_map = raw_fk_map
    else:
        fk_map = {}
    dropdown_ids = None
    if fk_map:
        from common.fk_resolver import FkResolver
        try:
            resolver = FkResolver(client)
            parent_screen = _screen_name(request.sub_module)
            dropdown_ids = {}
            for field, screen in fk_map.items():
                resolved = resolver.resolve(screen, parent_screen=parent_screen, field_key=field)
                if resolved:
                    dropdown_ids[field] = resolved
            # If item_ref_id was resolved, also fetch item→uom mapping from
            # Item Master so modules like CBR can use the correct UOM per item.
            if "item_ref_id" in dropdown_ids:
                try:
                    im_resp = client.list_entries("Item Master", page_size=500)
                    im_rows = im_resp.get("screenmatlistingdata_set", im_resp.get("results", im_resp.get("data", [])))
                    item_uom_map = {}
                    for row in (im_rows or []):
                        iid = row.get("id")
                        uom = row.get("uom") or row.get("base_uom")
                        if iid is not None and uom is not None:
                            item_uom_map[int(iid)] = uom
                    if item_uom_map:
                        dropdown_ids["item_uom_map"] = item_uom_map
                except Exception:
                    pass

            yield _sse_event(LogEvent(
                type="log",
                message=f"Resolved {len(dropdown_ids)} FK field(s) from ERP",
                timestamp=datetime.now(timezone.utc),
                run_id=run_id,
            ))
        except Exception as e:
            yield _sse_event(LogEvent(
                type="warning",
                message=f"FK resolution failed: {e}",
                timestamp=datetime.now(timezone.utc),
                run_id=run_id,
            ))

    # 4. Generate payloads
    start = time.time()
    created = 0
    failed = 0
    fill_all = (request.config or {}).get("fill_all", False)
    total = 9999 if fill_all else min(request.count, 500)
    records = []

    yield _sse_event(LogEvent(
        type="log",
        message=f"Generating {total} payloads...",
        timestamp=datetime.now(timezone.utc),
        run_id=run_id,
    ))

    # Query existing entries to avoid duplicates
    try:
        screen = _screen_name(request.sub_module)
        existing = client.list_entries(screen, page_size=500)
        existing_items = existing.get("screenmatlistingdata_set", existing.get("results", existing.get("data", [])))
        if isinstance(existing_items, list) and existing_items:
            yield _sse_event(LogEvent(
                type="log",
                message=f"Found {len(existing_items)} existing entries — will skip duplicates",
                timestamp=datetime.now(timezone.utc),
                run_id=run_id,
            ))
    except Exception:
        existing_items = []

    # If module needs item-level dedup, enrich listing rows with detail data
    enrich_fn = getattr(data_mod, "enrich_existing_entries", None)
    if enrich_fn and existing_items:
        try:
            existing_items = enrich_fn(client, existing_items)
        except Exception:
            pass

    if request.fixed_payloads:
        payloads = request.fixed_payloads
    else:
        sig = inspect.signature(generate_fn)
        kwargs = {"count": total}
        if dropdown_ids and "dropdown_ids" in sig.parameters:
            kwargs["dropdown_ids"] = dropdown_ids
        if existing_items and "existing_entries" in sig.parameters:
            kwargs["existing_entries"] = existing_items
        if "config" in sig.parameters:
            kwargs["config"] = request.config
        attr_number = (request.config or {}).get("attr_number", 1)
        if attr_number and "attr_number" in sig.parameters:
            kwargs["attr_number"] = attr_number
        payloads = generate_fn(**kwargs)

    # Item category override: if user selected a specific category from the UI, lock it in
    selected_item_category = (request.config or {}).get("item_category")
    if selected_item_category and request.sub_module == "item_master":
        for p in payloads:
            p["item_category"] = int(selected_item_category)

    # Conflict mode: override specific fields on every generated payload
    conflict_override = (request.config or {}).get("_conflict_override", {})
    if conflict_override and isinstance(conflict_override, dict):
        for p in payloads:
            p.update(conflict_override)

    # Tenant-universal FK pins (e.g. Customer Sale Type = "Commission", resolved live)
    pin_fn = getattr(data_mod, "apply_tenant_fk_pins", None)
    if callable(pin_fn):
        try:
            payloads = pin_fn(client, payloads)
        except Exception as e:
            yield _sse_event(LogEvent(
                type="warning",
                message=f"apply_tenant_fk_pins failed: {e}",
                timestamp=datetime.now(timezone.utc),
                run_id=run_id,
            ))
    batch_size = len(payloads)

    yield _sse_event(LogEvent(
        type="log",
        message=f"Created {batch_size} payloads (after dedup). Starting ERP API creation...",
        timestamp=datetime.now(timezone.utc),
        run_id=run_id,
    ))

    # Check if workflow transitions are requested
    workflow_cfg = (request.config or {}).get("workflow", {})
    do_verify = workflow_cfg.get("verify", False)
    do_approve = workflow_cfg.get("approve", False)

    skipped = 0

    for i, payload in enumerate(payloads):
        record_name = payload.get("name") or str(payload.get("attribute_name", ""))
        is_update = payload.get("id") and str(payload.get("id", "")).strip() not in ("", "0")
        result = client.update_entry(payload["id"], payload) if is_update else client.create_entry(payload)
        if result is not None:
            created += 1
            record_id = result.get("id", result.get("data", {}).get("id", "N/A"))
            action = "updated" if is_update else "created"
            records.append({"name": record_name, "record_id": record_id, "status": action})
            yield _sse_event(LogEvent(
                type="log",
                message=f"  [{i+1}/{total}] {action.upper()} #{record_id}  -  {record_name}",
                timestamp=datetime.now(timezone.utc),
                run_id=run_id,
            ))

            if do_verify or do_approve:
                try:
                    screen = _screen_name(request.sub_module)
                    verified = False
                    if do_verify:
                        v_result = client.workflow_action_from_entry(screen, record_id, "Verify")
                        if v_result is not None:
                            verified = True
                            records[-1]["status"] = "verified"
                            yield _sse_event(LogEvent(type="log", message=f"    -> #{record_id} verified", timestamp=datetime.now(timezone.utc), run_id=run_id))
                        else:
                            yield _sse_event(LogEvent(type="log", message=f"    -> #{record_id} verify FAILED ({_extract_wf_error(client)})", timestamp=datetime.now(timezone.utc), run_id=run_id))
                    if do_approve and (verified or not do_verify):
                        a_result = client.workflow_action_from_entry(screen, record_id, "Approve")
                        if a_result is not None:
                            records[-1]["status"] = "approved"
                            yield _sse_event(LogEvent(type="log", message=f"    -> #{record_id} approved", timestamp=datetime.now(timezone.utc), run_id=run_id))
                        else:
                            yield _sse_event(LogEvent(type="log", message=f"    -> #{record_id} approve FAILED ({_extract_wf_error(client)})", timestamp=datetime.now(timezone.utc), run_id=run_id))
                except Exception as wf_err:
                    yield _sse_event(LogEvent(type="log", message=f"    -> #{record_id} workflow error: {wf_err}", timestamp=datetime.now(timezone.utc), run_id=run_id))
        else:
            error_detail = getattr(client, '_last_raw_response', None)
            err_body = error_detail.text if error_detail is not None else ""
            err_status = error_detail.status_code if error_detail is not None else 0

            # Detect duplicate / already-exists errors → SKIPPED, not FAILED
            is_duplicate = err_status == 400 and any(
                kw in err_body.lower() for kw in ("duplicate", "already exist", "unique", "already present")
            )

            if is_duplicate:
                skipped += 1
                # Extract a friendly field value to show (e.g. uom_code)
                skip_label = (
                    payload.get("uom_code") or
                    payload.get("name") or
                    record_name
                )
                records.append({"name": record_name, "record_id": "-", "status": "skipped"})
                yield _sse_event(LogEvent(
                    type="log",
                    message=f"  [{i+1}/{total}] SKIPPED  -  {skip_label} (already exists)",
                    timestamp=datetime.now(timezone.utc),
                    run_id=run_id,
                ))
            else:
                failed += 1
                if err_status in (401, 403):
                    yield _sse_event(LogEvent(
                        type="auth_error",
                        message="ERP token is invalid or expired. Please update your token.",
                        timestamp=datetime.now(timezone.utc),
                        run_id=run_id,
                    ))
                err_msg = f"HTTP {err_status}: {err_body[:200]}" if error_detail is not None else "API returned None"
                records.append({"name": record_name, "record_id": "-", "status": "failed"})
                yield _sse_event(LogEvent(
                    type="log",
                    message=f"  [{i+1}/{total}]  FAILED  -  {record_name}: {err_msg}",
                    timestamp=datetime.now(timezone.utc),
                    run_id=run_id,
                ))

    elapsed = time.time() - start

    # 4. Save records to JSON for Excel export
    SUMMARY_FILE = RESULTS_DIR / f"{run_id}.json"
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    with open(SUMMARY_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "run_id": run_id,
            "module": request.module,
            "sub_module": request.sub_module,
            "total": total,
            "created": created,
            "failed": failed,
            "skipped": skipped,
            "elapsed_seconds": round(elapsed, 1),
            "records": records,
        }, f, indent=2, ensure_ascii=False)

    skipped_part = f", {skipped} skipped" if skipped else ""
    yield _sse_event(LogEvent(
        type="run_end",
        message=f"Batch create complete: {created} created{skipped_part}, {failed} failed in {elapsed:.1f}s",
        timestamp=datetime.now(timezone.utc),
        run_id=run_id,
    ))

    try:
        client.close()
    except Exception:
        pass


def export_batch_excel(run_id: str) -> str | None:
    """Generate a styled Excel file from saved batch results. Returns file path or None."""
    results_file = RESULTS_DIR / f"{run_id}.json"
    if not results_file.exists():
        return None

    with open(results_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Calibri", size=11)
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    fail_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Summary info
    ws.cell(row=1, column=1, value=f"Module: {data['module']} / {data['sub_module']}").font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws.cell(row=2, column=1, value=f"Total: {data['total']}  |  Created: {data['created']}  |  Failed: {data['failed']}  |  Duration: {data['elapsed_seconds']}s").font = Font(name="Calibri", size=10, color="555555")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

    # Header row
    headers = ["#", "Name", "Record ID"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for i, rec in enumerate(data["records"]):
        row = i + 5
        is_failed = rec["status"] == "failed"
        fill = fail_fill if is_failed else (alt_fill if i % 2 == 1 else None)

        ws.cell(row=row, column=1, value=i + 1).font = body_font
        ws.cell(row=row, column=2, value=rec["name"]).font = body_font
        ws.cell(row=row, column=3, value=rec["record_id"]).font = body_font

        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # Auto-filter
    ws.auto_filter.ref = f"A4:C{4 + len(data['records'])}"

    xlsx_path = RESULTS_DIR / f"{run_id}.xlsx"
    wb.save(xlsx_path)
    return str(xlsx_path)
