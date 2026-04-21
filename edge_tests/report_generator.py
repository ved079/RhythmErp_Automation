"""
Excel report generator for edge-case test results.

Usage:
    from edge_tests.report_generator import generate_report
    report_path = generate_report(results)
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Output directory — same as dashboard downloads
DOWNLOAD_DIR = os.path.join(
    os.path.abspath(os.path.join(os.path.dirname(__file__), '..')),
    "download_files", "Registration_TestCases"
)
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
PASS_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header(ws, row, cols):
    """Style a header row."""
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_cell(ws, row, col_idx, value, font=None, fill=None, align=None):
    """Style a data cell."""
    cell = ws.cell(row=row, column=col_idx, value=value)
    cell.font = font or NORMAL_FONT
    if fill:
        cell.fill = fill
    cell.alignment = align or LEFT
    cell.border = THIN_BORDER
    return cell


def _status_fill(status):
    """Return fill color based on status."""
    if status == "PASSED":
        return PASS_FILL
    elif status == "FAILED":
        return FAIL_FILL
    elif status == "ERROR":
        return ERROR_FILL
    return None


def generate_report(results):
    """
    Generate an Excel report from test results.

    Args:
        results: list of result dicts from runner.run_all_selected()

    Returns:
        str: absolute path to the generated Excel file
    """
    wb = Workbook()

    # ═══════════════════════════════════════════
    #  SHEET 1: SUMMARY
    # ═══════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Summary"

    _apply_header(ws_summary, 1, ["Test Case", "Total", "Passed", "Failed", "Pass Rate"])

    row = 2
    grand_total = 0
    grand_pass = 0
    grand_fail = 0

    for r in results:
        rate = f"{(r['pass_count'] / r['total'] * 100):.0f}%" if r['total'] > 0 else "N/A"
        fill = PASS_FILL if r['passed'] else FAIL_FILL

        _apply_cell(ws_summary, row, 1, r['test_case'], BOLD_FONT, fill, LEFT)
        _apply_cell(ws_summary, row, 2, r['total'], NORMAL_FONT, fill, CENTER)
        _apply_cell(ws_summary, row, 3, r['pass_count'], NORMAL_FONT, fill, CENTER)
        _apply_cell(ws_summary, row, 4, r['fail_count'], NORMAL_FONT, fill, CENTER)
        _apply_cell(ws_summary, row, 5, rate, NORMAL_FONT, fill, CENTER)

        grand_total += r['total']
        grand_pass += r['pass_count']
        grand_fail += r['fail_count']
        row += 1

    # Overall row
    overall_rate = f"{(grand_pass / grand_total * 100):.0f}%" if grand_total > 0 else "N/A"
    _apply_cell(ws_summary, row, 1, "OVERALL", BOLD_FONT, SUMMARY_FILL, LEFT)
    _apply_cell(ws_summary, row, 2, grand_total, BOLD_FONT, SUMMARY_FILL, CENTER)
    _apply_cell(ws_summary, row, 3, grand_pass, BOLD_FONT, SUMMARY_FILL, CENTER)
    _apply_cell(ws_summary, row, 4, grand_fail, BOLD_FONT, SUMMARY_FILL, CENTER)
    _apply_cell(ws_summary, row, 5, overall_rate, BOLD_FONT, SUMMARY_FILL, CENTER)

    # Run timestamp
    row += 2
    _apply_cell(ws_summary, row, 1, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}",
                Font(name="Calibri", size=9, italic=True, color="888888"), align=LEFT)

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 12

    # ═══════════════════════════════════════════
    #  SHEET 2: DETAILS
    # ═══════════════════════════════════════════
    ws_details = wb.create_sheet("Details")

    _apply_header(ws_details, 1, ["#", "Test Case", "Method", "Status", "Error Message", "Timestamp"])

    row = 2
    idx = 1
    for r in results:
        for detail in r['results']:
            fill = _status_fill(detail['status'])
            _apply_cell(ws_details, row, 1, idx, NORMAL_FONT, fill, CENTER)
            _apply_cell(ws_details, row, 2, r['test_case'], NORMAL_FONT, fill, LEFT)
            _apply_cell(ws_details, row, 3, detail['method'], NORMAL_FONT, fill, LEFT)
            _apply_cell(ws_details, row, 4, detail['status'], BOLD_FONT, fill, CENTER)
            _apply_cell(ws_details, row, 5, detail['error'] or "—", NORMAL_FONT, fill, LEFT)
            _apply_cell(ws_details, row, 6, detail['timestamp'], NORMAL_FONT, fill, CENTER)
            row += 1
            idx += 1

    ws_details.column_dimensions['A'].width = 5
    ws_details.column_dimensions['B'].width = 30
    ws_details.column_dimensions['C'].width = 45
    ws_details.column_dimensions['D'].width = 10
    ws_details.column_dimensions['E'].width = 60
    ws_details.column_dimensions['F'].width = 12

    # ═══════════════════════════════════════════
    #  SHEET 3: SCREENSHOTS (only failed tests)
    # ═══════════════════════════════════════════
    ws_shots = wb.create_sheet("Screenshots")

    _apply_header(ws_shots, 1, ["Test Case", "Method", "Screenshot Path"])

    row = 2
    has_screenshots = False
    for r in results:
        for detail in r['results']:
            if detail.get('screenshot'):
                has_screenshots = True
                fill = FAIL_FILL
                _apply_cell(ws_shots, row, 1, r['test_case'], NORMAL_FONT, fill, LEFT)
                _apply_cell(ws_shots, row, 2, detail['method'], NORMAL_FONT, fill, LEFT)
                _apply_cell(ws_shots, row, 3, detail['screenshot'], NORMAL_FONT, fill, LEFT)
                row += 1

    if not has_screenshots:
        _apply_cell(ws_shots, row, 1, "No failures — no screenshots captured.",
                    Font(name="Calibri", size=10, italic=True, color="27AE60"), align=LEFT)

    ws_shots.column_dimensions['A'].width = 30
    ws_shots.column_dimensions['B'].width = 45
    ws_shots.column_dimensions['C'].width = 60

    # ═══════════════════════════════════════════
    #  SAVE
    # ═══════════════════════════════════════════
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")

    if len(results) == 1:
        module_name = results[0]["test_case"].split()[0]  # "Farmer Registration TC" → "Farmer"
    else:
        module_name = "Registration"  # Multiple modules under same category
    filename = f"{module_name}_Test_Report_{timestamp}.xlsx"

    filepath = os.path.join(DOWNLOAD_DIR, filename)

    wb.save(filepath)
    return os.path.abspath(filepath)
