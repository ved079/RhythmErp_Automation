from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import sys
import os
import queue
import threading
import uuid
import asyncio
import logging
import time
import traceback
import pathlib
from datetime import datetime

import openpyxl

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, project_root)

import config
from common import auth_section, nav_section

from privateb2b.test_cases import stock_transfer_recon
from privateb2b import stock_transfer_section

from privateb2b.purchase import (
    gatepass_section, grn_section,
    purchase_booking_section, qc_section
)
from privateb2b.sales import (
    dispatch_note_section, invoice_section,
    lot_creation_section, receipt_section,
    sales_order_section
)

from Registration import (
    farmer_section,
    supplier_section,
    agent_section,
    customer_section,
    employee_section
)

from data.test_data import (
    farmer_data,
    supplier_data,
    agent_data,
    customer_data,
    employee_data,
    gatepass_data,
    grn_data,
    qc_data,
    purchase_booking_data,
    sales_order_data,
    dispatch_note_data,
    invoice_data,
    receipt_data,
    stock_transfer_data
)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Serve Dashboard ----------
@app.get("/")
async def serve_dashboard():
    html_path = pathlib.Path(__file__).parent / "index.html"
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))

# ---------- Download directory (dynamic path, scans recursively) ----------
DOWNLOAD_DIR = os.path.join(project_root, "download_files")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# ---------- Models ----------
class LoginCredentials(BaseModel):
    username: str
    password: str
    tenant: str

class AutomationRequest(BaseModel):
    username: str
    password: str
    tenant: str
    modules: list[str]

# ---------- Store for active executions ----------
active_executions = {}

# ---------- Custom logging handler ----------
class QueueHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        self.log_queue.put(self.format(record))

# ---------- Session keep-alive ----------
def session_keepalive(driver, stop_event, interval=120):
    """Keeps the ERP session alive by refreshing the page every N seconds."""
    while not stop_event.is_set():
        stop_event.wait(interval)
        if not stop_event.is_set():
            try:
                driver.refresh()
                time.sleep(3)
            except Exception:
                break

# ---------- Modal cleanup helper ----------
def close_open_modals(driver, wait):
    """Closes any open modals/overlays left behind by previous modules."""
    try:
        close_selectors = [
            "//button[contains(text(),'Close')]",
            "//button[contains(@class,'close')]",
            "//button[@aria-label='Close']",
            "//span[contains(@class,'close')]",
            "//div[contains(@class,'modal')]//button[contains(@class,'btn')]",
            "//div[@role='dialog']//button",
        ]
        for selector in close_selectors:
            try:
                buttons = driver.find_elements(By.XPATH, selector)
                for btn in buttons:
                    try:
                        if btn.is_displayed() and btn.is_enabled():
                            driver.execute_script("arguments[0].click();", btn)
                            time.sleep(0.5)
                            return True
                    except Exception:
                        continue
            except Exception:
                continue

        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(0.5)
        except Exception:
            pass

        try:
            backdrop = driver.find_element(By.XPATH,
                "//div[contains(@class,'modal-backdrop') or contains(@class,'overlay')]")
            if backdrop.is_displayed():
                driver.execute_script(
                    "arguments[0].click();", backdrop)
                time.sleep(0.5)
        except Exception:
            pass

        return True
    except Exception:
        return False

# ---------- Automation worker ----------
def automation_worker(exec_id: str, req: AutomationRequest, log_queue: queue.Queue, command_queue: queue.Queue):
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    handler = QueueHandler(log_queue)
    handler.setFormatter(logging.Formatter('%(asctime)s | %(message)s', datefmt='%H:%M:%S'))
    logger.addHandler(handler)

    driver = None
    stop_event = threading.Event()
    screenshot_thread = None
    command_thread = None
    keepalive_thread = None
    keepalive_event = threading.Event()
    browser_visible = False

    def screenshot_loop():
        """Captures screenshots every 1 second."""
        while not stop_event.is_set():
            try:
                screenshot_b64 = driver.get_screenshot_as_base64()
                log_queue.put("SCREENSHOT:data:image/png;base64," + screenshot_b64)
            except Exception:
                break
            stop_event.wait(1.0)

    def command_listener():
        """Dedicated thread that responds to toggle commands instantly."""
        nonlocal browser_visible
        while not stop_event.is_set():
            try:
                cmd = command_queue.get(timeout=0.5)
                if cmd == "TOGGLE":
                    browser_visible = not browser_visible
                    try:
                        if browser_visible:
                            driver.set_window_position(100, 100)
                            driver.maximize_window()
                            log_queue.put("UI_TRIGGER:BROWSER_SHOW")
                            logging.info("Browser window shown (visible)")
                        else:
                            driver.set_window_position(-2000, -2000)
                            log_queue.put("UI_TRIGGER:BROWSER_HIDE")
                            logging.info("Browser window hidden (off-screen)")
                    except Exception as e:
                        logging.error(f"Failed to toggle browser: {e}")
                        log_queue.put(f"ERROR: Toggle failed - {e}")
                elif cmd == "STOP":
                    break
            except queue.Empty:
                continue
            except Exception as e:
                logging.error(f"Command listener error: {e}")
                break

    try:
        log_queue.put("UI_TRIGGER:AUTH_START")
        logging.info("Starting Chrome...")
        options = webdriver.ChromeOptions()
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        # Move Chrome off-screen immediately
        driver.set_window_position(-2000, -2000)
        wait = WebDriverWait(driver, 60)

        # Start screenshot thread
        screenshot_thread = threading.Thread(target=screenshot_loop, daemon=True)
        screenshot_thread.start()
        # Start command listener thread
        command_thread = threading.Thread(target=command_listener, daemon=True)
        command_thread.start()

        # Start session keep-alive (pings every 2 minutes)
        keepalive_thread = threading.Thread(
            target=session_keepalive,
            args=(driver, keepalive_event, 120),
            daemon=True
        )
        keepalive_thread.start()
        logging.info("Session keep-alive started (pings every 2 min).")

        # Override config
        original = (config.URL, config.USER, config.PASS, config.TENANT_NAME)
        config.USER = req.username
        config.PASS = req.password
        config.TENANT_NAME = req.tenant

        logging.info(f"Logging in with {req.username} @ {req.tenant}")
        auth_section.perform_login(driver, wait, config)
        logging.info("Login Successful")
        log_queue.put("UI_TRIGGER:AUTH_DONE")
        log_queue.put("UI_TRIGGER:EXEC_START")

        # ───────────────────────────────────────────
        #  Separate TC modules from regular modules
        # ───────────────────────────────────────────
        tc_modules = [m for m in req.modules if m.endswith(" TC")]
        regular_modules = [m for m in req.modules if not m.endswith(" TC")]

        # ========== REGISTRATION MODULES ==========
        if "Farmer Registration" in regular_modules:
            logging.info("Running Farmer Registration...")
            nav_section.go_to_farmer_page(driver, wait)
            farmer_section.fill_registration(driver, wait, farmer_data)
            logging.info("Farmer Registration complete.")

        if "Supplier Registration" in regular_modules:
            logging.info("Running Supplier Registration...")
            nav_section.go_to_supplier_page(driver, wait)
            supplier_section.fill_supplier_registration(driver, wait, supplier_data)
            logging.info("Supplier Registration complete.")

        if "Agent Registration" in regular_modules:
            logging.info("Running Agent Registration...")
            nav_section.go_to_agent_page(driver, wait)
            agent_section.fill_agent_registration(driver, wait, agent_data)
            logging.info("Agent Registration complete.")

        if "Customer Registration" in regular_modules:
            logging.info("Running Customer Registration...")
            nav_section.go_to_customer_page(driver, wait)
            customer_section.fill_customer_registration(driver, wait, customer_data)
            logging.info("Customer Registration complete.")

        if "Employee Registration" in regular_modules:
            logging.info("Running Employee Registration...")
            nav_section.go_to_employee_page(driver, wait)
            employee_section.fill_employee_registration(driver, wait, employee_data)
            logging.info("Employee Registration complete.")

        # ========== PURCHASE FLOW ==========
        if "Purchase Flow" in regular_modules:
            logging.info("Running Purchase Flow...")

            # Gatepass
            nav_section.go_to_gatepass_page(driver, wait)
            gatepass_section.fill_gatepass_registration(driver, wait, gatepass_data)
            logging.info("Gatepass completed.")
            close_open_modals(driver, wait)

            # GRN
            # nav_section.go_to_grn_page(driver, wait)
            # grn_section.fill_grn_registration(driver, wait, grn_data)
            # time.sleep(10)
            # grn_section.approve_latest_grn(driver, wait)
            # logging.info("GRN completed.")
            # close_open_modals(driver, wait)

            # QC
            nav_section.go_to_qc_page(driver, wait)
            qc_section.fill_qc_registration(driver, wait, qc_data)
            logging.info("QC completed.")
            close_open_modals(driver, wait)

            # Purchase Booking
            nav_section.go_to_purchase_booking_page(driver, wait)
            purchase_booking_section.fill_purchase_booking_registration(driver, wait, purchase_booking_data)
            logging.info("Purchase Booking completed.")

            # Close the View modal left open by search_and_export_latest_pb
            close_open_modals(driver, wait)
            logging.info("Purchase Flow complete.")

        # ========== SALES FLOW ==========
        if "Sales Flow" in regular_modules:
            logging.info("Running Sales Flow...")

            nav_section.go_to_sales_order_page(driver, wait)
            sales_order_section.fill_sales_order_registration(driver, wait, sales_order_data)
            logging.info("Sales Order completed.")
            close_open_modals(driver, wait)

            nav_section.go_to_lot_creation_page(driver, wait)
            lot_creation_section.fill_lot_creation(driver, wait, {
                "customer_name": sales_order_data['customer_name'],
                "items": sales_order_data['items']
            })
            logging.info("Lot Creation completed.")
            close_open_modals(driver, wait)

            nav_section.go_to_dispatch_note_page(driver, wait)
            dispatch_note_section.fill_dispatch_note_registration(driver, wait, dispatch_note_data)
            logging.info("Dispatch Note completed.")
            close_open_modals(driver, wait)

            nav_section.go_to_invoice_page(driver, wait)
            invoice_section.fill_invoice_registration(driver, wait, invoice_data)
            logging.info("Invoice completed.")
            close_open_modals(driver, wait)

            nav_section.go_to_receipt_page(driver, wait)
            receipt_section.fill_receipt_registration(driver, wait, receipt_data)
            logging.info("Receipt completed.")

            logging.info("Sales Flow complete.")

        # ========== STOCK TRANSFER RECONCILIATION ==========
        if "Stock Transfer Reconciliation" in regular_modules:
            logging.info("Running Stock Transfer Reconciliation...")

            nav_section.go_to_stock_transfer_page(driver, wait)
            stock_transfer_section.fill_stock_transfer(driver, wait, stock_transfer_data)

            logging.info("Stock Transfer Complete.")
            time.sleep(2)

        # ========== AGEING REPORT ==========
        if "Ageing Report" in regular_modules:
            logging.info("Running Ageing Report...")
            time.sleep(2)

        # ========== DOWNLOADS ==========
        if "Downloaded Recon Files" in regular_modules:
            logging.info("Opening file browser in dashboard...")
            log_queue.put("UI_TRIGGER:OPEN_FILES")
            logging.info("Files view triggered.")

        # ========== EDGE-CASE TEST MODULES ==========
        tc_results = None
        if tc_modules:
            logging.info(f"Running {len(tc_modules)} test case module(s)...")
            log_queue.put(f"UI_TRIGGER:EXEC_TEXT:Running TCs: {', '.join(tc_modules)}")

            # PAUSE keepalive — it refreshes the page and kills mid-test state
            keepalive_event.set()
            logging.info("Keepalive paused for test case execution.")

            try:
                from edge_tests.runner import run_all_selected, generate_report as gen_tc_report
                tc_results = run_all_selected(tc_modules, driver, wait, skip_login=True)


                # Print summary to terminal

                grand_total = 0
                grand_pass = 0
                grand_fail = 0
                for r in tc_results:
                    logging.info(f"  {r['test_case']}: {r['summary']}")
                    for d in r['results']:
                        icon = "PASSED" if d['status'] == "PASSED" else "FAILED"
                        logging.info(f"    {icon} - {d['method']}")
                    grand_total += r['total']
                    grand_pass += r['pass_count']
                    grand_fail += r['fail_count']

                rate = f"{(grand_pass / grand_total * 100):.0f}%" if grand_total > 0 else "N/A"
                logging.info(f"Test Cases Complete: {grand_pass}/{grand_total} passed ({rate})")

            except ImportError as e:
                logging.error(f"Could not import test runner: {e}")
                log_queue.put(f"ERROR: Test runner not available - {e}")
            except Exception as e:
                logging.error(f"Test execution failed: {e}")
                log_queue.put(f"ERROR: Test execution failed - {e}")
            finally:
                # RESUME keepalive after TC runs
                keepalive_event.clear()
                logging.info("Keepalive resumed.")

        # ========== REPORT GENERATION ==========
        log_queue.put("UI_TRIGGER:EXEC_DONE")
        log_queue.put("UI_TRIGGER:REPORT_START")

        if tc_results:
            try:
                from edge_tests.runner import generate_report as gen_tc_report
                logging.info("Generating test case Excel report...")
                report_path = gen_tc_report(tc_results)
                if report_path:
                    rel_path = os.path.relpath(report_path, DOWNLOAD_DIR)
                    log_queue.put(f"UI_TRIGGER:REPORT_FILE:{rel_path}")
                    logging.info(f"Report saved: {report_path}")
                else:
                    logging.warning("Report generation returned no path")
            except Exception as e:
                logging.error(f"Report generation failed: {e}")
                log_queue.put(f"ERROR: Report generation failed - {e}")
        else:
            logging.info("No test cases selected. Skipping report generation.")
            time.sleep(1)

        logging.info("Reports ready.")
        log_queue.put("UI_TRIGGER:REPORT_DONE")

        # Restore config
        (config.URL, config.USER, config.PASS, config.TENANT_NAME) = original

    except Exception as e:
        error_trace = traceback.format_exc()
        logging.error(f"Automation failed: {e}\n{error_trace}")
        log_queue.put(f"ERROR: {e}")
        # ALWAYS send completion triggers on error so the frontend unlocks
        log_queue.put("UI_TRIGGER:EXEC_DONE")
        log_queue.put("UI_TRIGGER:REPORT_DONE")

    finally:
        # 1. Signal all helper threads to stop
        stop_event.set()
        keepalive_event.set()
        try:
            command_queue.put_nowait("STOP")
        except queue.Full:
            pass

        # 2. Join helper threads
        for t in [screenshot_thread, command_thread, keepalive_thread]:
            if t and t.is_alive():
                t.join(timeout=3)

        # 3. Quit driver safely
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

        # 4. Remove logging handler
        logger.removeHandler(handler)

        # 5. ALWAYS send ALL_DONE
        log_queue.put("UI_TRIGGER:ALL_DONE")


# ---------- Endpoints ----------
@app.post("/validate_login")
def validate_login(creds: LoginCredentials):
    driver = None
    try:
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        wait = WebDriverWait(driver, 30)

        original = (config.USER, config.PASS, config.TENANT_NAME)
        config.USER = creds.username
        config.PASS = creds.password
        config.TENANT_NAME = creds.tenant

        auth_section.perform_login(driver, wait, config)

        (config.USER, config.PASS, config.TENANT_NAME) = original
        return {"status": "success", "message": "Login validated successfully"}
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid credentials or ERP unreachable")
    finally:
        if driver:
            driver.quit()


@app.post("/run_automation")
async def run_automation(req: AutomationRequest):
    exec_id = str(uuid.uuid4())
    log_queue = queue.Queue()
    command_queue = queue.Queue()
    thread = threading.Thread(target=automation_worker, args=(exec_id, req, log_queue, command_queue))
    thread.start()
    active_executions[exec_id] = (thread, log_queue, command_queue)
    return {"execution_id": exec_id}


@app.post("/toggle_browser/{exec_id}")
async def toggle_browser(exec_id: str):
    """Toggle Chrome between off-screen and visible."""
    if exec_id not in active_executions:
        raise HTTPException(status_code=404, detail="Execution not found")
    _, _, command_queue = active_executions[exec_id]
    command_queue.put("TOGGLE")
    return {"status": "ok"}


# ---------- File Browser Endpoints (recursive) ----------
@app.get("/api/files/list")
def list_files():
    """
    Recursively list all files inside download_files/.
    Returns each file with its relative path (e.g. "Registration_TestCases/report.xlsx").
    """
    files = []
    if not os.path.isdir(DOWNLOAD_DIR):
        return {"files": files}

    for root, dirs, filenames in os.walk(DOWNLOAD_DIR):
        for fname in filenames:
            full_path = os.path.join(root, fname)
            rel_path = os.path.relpath(full_path, DOWNLOAD_DIR)
            stat = os.stat(full_path)
            mtime = datetime.fromtimestamp(stat.st_mtime).strftime("%d-%b-%Y %H:%M")
            files.append({
                "name": rel_path.replace("\\", "/"),
                "size": stat.st_size,
                "modified": mtime,
            })

    # Sort newest first
    files.sort(key=lambda f: f["modified"], reverse=True)
    return {"files": files}


@app.get("/api/files/read/{file_path:path}")
def read_excel_file(file_path: str):
    """
    Read an Excel file from download_files/ and return its sheets as JSON.
    file_path is relative to DOWNLOAD_DIR, e.g. "Registration_TestCases/report.xlsx"
    """
    full_path = os.path.join(DOWNLOAD_DIR, file_path.replace("/", os.sep))

    if not os.path.isfile(full_path):
        raise HTTPException(status_code=404, detail="File not found")

    try:
        wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell) if cell is not None else "" for cell in row]
                else:
                    if any(cell is not None for cell in row):
                        rows.append([str(cell) if cell is not None else "" for cell in row])
            if headers:
                sheets[sheet_name] = {"headers": headers, "rows": rows}
        wb.close()
        return {"sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")


@app.get("/api/files/download/{file_path:path}")
def download_file(file_path: str):
    """
    Serve a file from download_files/ for browser download.
    file_path is relative to DOWNLOAD_DIR.
    """
    full_path = os.path.join(DOWNLOAD_DIR, file_path.replace("/", os.sep))

    if not os.path.isfile(full_path):
        raise HTTPException(status_code=404, detail="File not found")

    filename = os.path.basename(full_path)
    return FileResponse(
        path=full_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.websocket("/ws/logs/{exec_id}")
async def websocket_logs(websocket: WebSocket, exec_id: str):
    await websocket.accept()
    if exec_id not in active_executions:
        await websocket.send_text("ERROR: Invalid execution ID")
        await websocket.close()
        return

    _, log_queue, _ = active_executions[exec_id]
    try:
        while True:
            try:
                msg = log_queue.get(timeout=0.1)
                await websocket.send_text(msg)
                if msg == "UI_TRIGGER:ALL_DONE":
                    break
            except queue.Empty:
                if exec_id not in active_executions:
                    while True:
                        try:
                            msg = log_queue.get_nowait()
                            await websocket.send_text(msg)
                            if msg == "UI_TRIGGER:ALL_DONE":
                                break
                        except queue.Empty:
                            break
                    break
                await asyncio.sleep(0.05)
    except WebSocketDisconnect:
        pass
    finally:
        active_executions.pop(exec_id, None)
