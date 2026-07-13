"""
Excel exporter for Purchase Flow integration tests.
Generates a timestamped .xlsx report in the reports/ folder next to this file.

Usage (call at the end of a test class or conftest teardown):
    from pages.private_b2b.modules.Purchase_Flow_Tests.excel_exporter import export_pogpgrn_flow
    export_pogpgrn_flow(integration_state, step_results)
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")

# ── Colours ──────────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL  = PatternFill("solid", fgColor="FFEB9C")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")

WHITE_FONT  = Font(bold=True, color="FFFFFF")
BOLD_FONT   = Font(bold=True)
TITLE_FONT  = Font(bold=True, size=14, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_row(ws, row, cols, fill=HEADER_FILL, font=WHITE_FONT):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _data_cell(ws, row, col, value, fill=None, bold=False, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if bold:
        cell.font = BOLD_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = THIN_BORDER
    return cell


def _match_fill(expected, got, tolerance=0.5):
    if expected is None or got is None:
        return AMBER_FILL
    try:
        return GREEN_FILL if abs(float(expected) - float(got)) <= tolerance else RED_FILL
    except (TypeError, ValueError):
        return GREEN_FILL if str(expected).strip() == str(got).strip() else RED_FILL


def _match_symbol(expected, got, tolerance=0.5):
    fill = _match_fill(expected, got, tolerance)
    return "✓" if fill == GREEN_FILL else ("?" if fill == AMBER_FILL else "✗")


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_flow_trace(wb, state):
    ws = wb.create_sheet("Flow Trace")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Flow Trace — PO → GP → GRN"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Run at: {state.get('run_timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}"
    ts_cell.alignment = Alignment(horizontal="center")

    _header_row(ws, 4, ["Step", "Ref No", "Supplier", "Item(s)", "Location", "Qty / Bags", "Rate", "Status"])

    rows = [
        (
            "Purchase Order",
            state.get("po_ref_no", "—"),
            state.get("supplier_name", "—"),
            ", ".join(state.get("po_item_rates", {}).keys()) or "—",
            state.get("location", "—"),
            "500 each",
            "—",
            "CREATED" if state.get("po_ref_no") else "FAILED",
        ),
        (
            "Gate Pass",
            state.get("gp_ref_no", "—"),
            state.get("supplier_name", "—"),
            state.get("gp_item_name", "—"),
            state.get("location", "—"),
            f"{state.get('gp_qty', '—')} ({state.get('gp_bags', '—')} bags)",
            "—",
            "CREATED" if state.get("gp_ref_no") else "FAILED",
        ),
        (
            "GRN",
            state.get("grn_ref_no", "—"),
            state.get("supplier_name", "—"),
            state.get("gp_item_name", "—"),
            state.get("location", "—"),
            state.get("grn_accepted_qty", "—"),
            state.get("grn_rate", "—"),
            "CREATED" if state.get("grn_ref_no") else "FAILED",
        ),
    ]

    for r, row_data in enumerate(rows, 5):
        status = row_data[-1]
        row_fill = GREEN_FILL if status == "CREATED" else RED_FILL
        for c, val in enumerate(row_data, 1):
            cell_fill = row_fill if c == 8 else None
            _data_cell(ws, r, c, val, fill=cell_fill, align="left" if c in (2, 3, 4) else "center")

    _auto_width(ws)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[4].height = 20
    return ws


def _build_qty_check(wb, state):
    ws = wb.create_sheet("Qty & Rate Check")

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Quantity & Rate Verification"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    _header_row(ws, 3, ["Check", "Item", "Expected", "Got", "Diff", "Match", "Notes"])

    gp_rows  = state.get("gp_rows", [])
    po_rates = state.get("po_item_rates", {})

    checks = []

    # PO → GP qty check (per row)
    for gp_row in gp_rows:
        item   = gp_row.get("item_name", "—")
        gp_qty = gp_row.get("qty")
        checks.append({
            "check":    "GP qty vs GRN gate_pass_qty",
            "item":     item,
            "expected": gp_qty,
            "got":      state.get("grn_gate_pass_qtys", {}).get(item, gp_qty),
            "notes":    "GRN should auto-fill GP qty",
        })
        # Rate check
        po_rate  = po_rates.get(item)
        grn_rate = state.get("grn_rates", {}).get(item)
        if po_rate is not None:
            checks.append({
                "check":    "Rate: PO vs GRN",
                "item":     item,
                "expected": po_rate,
                "got":      grn_rate,
                "notes":    "GRN rate must match PO rate",
            })
        # PO remaining vs GP qty
        po_remaining = state.get("grn_po_remaining", {}).get(item)
        if po_remaining is not None and gp_qty is not None:
            checks.append({
                "check":    "PO remaining ≥ GP qty",
                "item":     item,
                "expected": f"≥ {gp_qty}",
                "got":      po_remaining,
                "notes":    "PO must have enough balance",
            })

    # GRN accepted qty check
    grn_accepted = state.get("grn_accepted_qty")
    if grn_accepted is not None and gp_rows:
        total_gp = sum(r.get("qty", 0) for r in gp_rows)
        checks.append({
            "check":    "GRN accepted qty vs GP total",
            "item":     "All items",
            "expected": total_gp,
            "got":      grn_accepted,
            "notes":    "Full GP qty accepted",
        })

    for r, chk in enumerate(checks, 4):
        exp = chk["expected"]
        got = chk["got"]
        # For "≥ N" style checks handle separately
        if isinstance(exp, str) and exp.startswith("≥"):
            threshold = float(exp[1:].strip())
            match = got is not None and float(got) >= threshold
            symbol = "✓" if match else "✗"
            fill   = GREEN_FILL if match else RED_FILL
            diff   = float(got) - threshold if got is not None else "?"
        else:
            symbol = _match_symbol(exp, got)
            fill   = _match_fill(exp, got)
            try:
                diff = round(float(got or 0) - float(exp or 0), 4)
            except (TypeError, ValueError):
                diff = "—"

        _data_cell(ws, r, 1, chk["check"],  align="left")
        _data_cell(ws, r, 2, chk["item"],   align="left")
        _data_cell(ws, r, 3, exp)
        _data_cell(ws, r, 4, got)
        _data_cell(ws, r, 5, diff)
        _data_cell(ws, r, 6, symbol, fill=fill, bold=True)
        _data_cell(ws, r, 7, chk.get("notes", ""), align="left")

    _auto_width(ws)
    return ws


def _build_run_summary(wb, state, step_results):
    ws = wb.create_sheet("Run Summary")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Run Summary — TestPOGPGRNFlow"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    _header_row(ws, 3, ["Step", "Test Method", "Status", "Duration (s)", "Ref Created", "Notes"])

    steps = [
        ("Step 1", "test_step1_create_po",         state.get("po_ref_no")),
        ("Step 2", "test_step2_create_gp",         state.get("gp_ref_no")),
        ("Step 3", "test_step3_create_grn_and_verify", state.get("grn_ref_no")),
    ]

    passed = 0
    failed = 0
    for r, (step, method, ref) in enumerate(steps, 4):
        result = step_results.get(method, {})
        status   = result.get("status", "PASSED" if ref else "FAILED")
        duration = result.get("duration_s", "—")
        notes    = result.get("notes", "")
        fill     = GREEN_FILL if status == "PASSED" else RED_FILL
        if status == "PASSED":
            passed += 1
        else:
            failed += 1

        _data_cell(ws, r, 1, step)
        _data_cell(ws, r, 2, method, align="left")
        _data_cell(ws, r, 3, status, fill=fill, bold=True)
        _data_cell(ws, r, 4, duration)
        _data_cell(ws, r, 5, ref or "—")
        _data_cell(ws, r, 6, notes, align="left")

    # Totals row
    total_row = len(steps) + 4
    _data_cell(ws, total_row, 1, "TOTAL", bold=True)
    _data_cell(ws, total_row, 3, f"{passed} passed / {failed} failed",
               fill=GREEN_FILL if failed == 0 else RED_FILL, bold=True)
    _data_cell(ws, total_row, 5, state.get("run_timestamp", ""))

    _auto_width(ws)
    return ws


# ── Public API ────────────────────────────────────────────────────────────────

def export_pogpgrn_flow(integration_state: dict, step_results: dict | None = None) -> str:
    """
    Generate Excel report for TestPOGPGRNFlow.

    Args:
        integration_state: the class-scoped dict populated by test steps.
        step_results: optional dict keyed by test method name with keys
                      'status' ('PASSED'/'FAILED'), 'duration_s', 'notes'.

    Returns:
        Absolute path to the saved .xlsx file.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)

    ts  = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    integration_state.setdefault("run_timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _build_run_summary(wb, integration_state, step_results or {})
    _build_flow_trace(wb, integration_state)
    _build_qty_check(wb, integration_state)

    path = os.path.join(REPORTS_DIR, f"TestPOGPGRNFlow_{ts}.xlsx")
    wb.save(path)
    print(f"\n[EXCEL] Report saved -> {path}")
    return path
