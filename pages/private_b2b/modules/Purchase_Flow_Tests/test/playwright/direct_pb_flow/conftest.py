"""
Conftest for Direct PB creation tests.

Login: kedar@rhythmflows.com / Kedar@999999 -> tenant: Eco Green Pvt Ltd.
Isolated from all other test modules — separate browser and session.
"""

import datetime
import os
import sys
import pytest
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.sync_api import sync_playwright




PROJECT_ROOT = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", "..", "..", "..")
)
sys.path.insert(0, PROJECT_ROOT)

from pages.private_b2b.modules.purchase_booking.direct_pb_playwright_page import DirectPBPlaywrightPage

LOGIN_URL = os.environ.get("RHYTHMERP_LOGIN_URL", "https://rhythmerp.algorhythms.in")
EMAIL     = os.environ.get("RHYTHMERP_DIRECT_PB_EMAIL",    "kedar@rhythmflows.com")
PASSWORD  = os.environ.get("RHYTHMERP_DIRECT_PB_PASSWORD", "Kedar@999999")
TENANT    = os.environ.get("RHYTHMERP_DIRECT_PB_TENANT",   "Eco Green Pvt Ltd")


def pytest_configure(config):
    config.addinivalue_line("markers", "direct_pb: Direct PB creation tests")


# ── Excel report ──────────────────────────────────────────────────────────────
_REPORT_META = [
    # (class, test_name, what_it_verifies, run_cmd_suffix)
    ("TestRowAndPostSaveSuite",    "test_rps1_row_behaviour_and_post_save",              "Row add/delete/edit behaviour + save verification",              "::TestRowAndPostSaveSuite"),
    ("TestDirectPBCalculations",   "test_cs1_all_single_row_calc_checks",                "All single-row calc checks (amount, disc, GST, total)",          "::TestDirectPBCalculations"),
    ("",                           "test_c_tc1_amount_auto_computed_from_rate_and_net_qty", "Amount = rate × net_qty",                                     ""),
    ("",                           "test_c_tc2_discount_amount_equals_amount_times_disc_pct", "Disc amount = amount × disc%",                              ""),
    ("",                           "test_c_tc3_total_no_tax_equals_amount_minus_discount",   "Total (no tax) = amount − discount",                         ""),
    ("",                           "test_c_tc4_total_deducts_labour_charges",                "Total deducts labour",                                       ""),
    ("",                           "test_c_tc5_net_qty_reduces_by_empty_bag_weight",         "net_qty = qty − EBW",                                        ""),
    ("",                           "test_c_tc6_net_qty_equals_full_qty_when_no_ebw",         "net_qty = qty when EBW=0",                                   ""),
    ("",                           "test_c_tc7_fractional_ebw_gives_fractional_net_qty",     "Fractional EBW → fractional net_qty",                        ""),
    ("",                           "test_c_tc8_igst_amount_equals_amount_times_tax_rate",    "IGST = amount × tax rate",                                   ""),
    ("",                           "test_c_tc9_cgst_and_sgst_each_half_of_igst_equivalent",  "CGST = SGST = IGST / 2",                                     ""),
    ("",                           "test_c_tc10_header_total_equals_sum_of_two_rows",        "Header total = sum of 2 rows",                               ""),
    ("TestMultiRowCalculations",   "test_m1_five_rows_varied_inputs_header_equals_sum",      "5 rows varied inputs → header = sum",                        "::TestMultiRowCalculations"),
    ("",                           "test_m1b_header_disc_equals_sum_of_row_discs",           "Header discount = sum of row discounts",                     ""),
    ("",                           "test_m1c_ebw_per_row_net_qty_affects_totals",            "EBW per row → net_qty affects totals",                       ""),
    ("",                           "test_m1d_all_20_items_header_equals_sum",                "20 rows → header = sum",                                     ""),
    ("",                           "test_m1f_many_rows_all_fields_random",                   "Many rows, all fields random → header = sum",                ""),
    ("",                           "test_m1e_gst_mixed_all_rows_header_sum",                 "GST mixed across rows → header sum",                         ""),
    ("",                           "test_m1g_gst_split_verification",                        "CGST/SGST split correctness",                                ""),
    ("",                           "test_m1h_gst_validation_future_ready",                   "GST validation (future-ready assertions)",                   ""),
    ("TestValidationSuite",        "test_vs1_all_validations_one_form",                      "All PB form validations in one form session",                 "::TestValidationSuite"),
    ("TestRowMutations",           "test_m2_add_n_delete_random_header_updates",             "Add N rows then delete random → header updates",             "::TestRowMutations"),
    ("",                           "test_m2b_delete_all_but_one_header_equals_single_row",   "Delete all but 1 → header = remaining row",                  ""),
    ("",                           "test_m3_mutate_disc_delete_add_recheck",                 "Mutate disc, delete, add → header recalcs",                  ""),
    ("",                           "test_m3b_change_qty_on_existing_rows_header_recalcs",    "Change qty on existing rows → header recalcs",               ""),
    ("",                           "test_m3c_change_disc_on_existing_rows_header_recalcs",   "Change disc on existing rows → header recalcs",              ""),
    ("TestEdgeSuite",              "test_es1_all_edge_cases",                                "21 edge case scenarios (EBW=qty, 100% disc, GST switch, etc.)", "::TestEdgeSuite"),
    ("TestDecimalPrecision",       "test_dp1_5dp_qty_rejected_4dp_accepted_total_2dp",       "5dp qty rejected, 4dp accepted, total rounded to 2dp",       "::TestDecimalPrecision"),
    ("TestUserBehaviourSuite",     "test_ub_tc1_fill_details_before_item_multirow",          "Fill qty/disc/labour first, item last — calcs fire correctly","::TestUserBehaviourSuite"),
    ("",                           "test_ub_tc2_fill_ebw_before_item",                       "EBW filled before item — net_qty = qty − EBW after item selected", ""),
    ("",                           "test_ub_tc3_swap_item_after_fields_filled",              "Swap item after fields filled — disc% retained, total recalculates", ""),
    ("",                           "test_ub_tc4_delete_bottom_up",                           "Delete rows bottom-up — header correct at every step",       ""),
    ("",                           "test_ub_tc5_delete_middle_rows",                         "Delete middle rows — header = first + last row",             ""),
    ("",                           "test_ub_tc6_delete_readd_same_item",                     "Delete item row and re-add same item — fresh calc, no stale data", ""),
    ("",                           "test_ub_tc7_sequential_edits_across_rows",               "Sequential edits across rows — header tracks all changes",   ""),
]

_BASE_CMD = 'python -m pytest "pages/private_b2b/modules/Purchase_Flow_Tests/test/playwright/direct_pb_flow/test_direct_pb_flow.py'

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")
_GREY_FILL = PatternFill("solid", fgColor="D9D9D9")


def _col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def pytest_terminal_summary(terminalreporter, exitstatus, config):
    passed  = {r.nodeid.split("::")[-1] for r in terminalreporter.stats.get("passed",  [])}
    failed  = {r.nodeid.split("::")[-1] for r in terminalreporter.stats.get("failed",  [])}
    skipped = {r.nodeid.split("::")[-1] for r in terminalreporter.stats.get("skipped", [])}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Direct PB Test Report"

    headers = ["#", "Class", "Test Name", "What it verifies", "Ran?", "Passed?", "Run Command"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER

    row_num = 2
    for idx, (cls, test_name, what, cmd_suffix) in enumerate(_REPORT_META, 1):
        ran    = test_name in passed or test_name in failed or test_name in skipped
        passed_ = test_name in passed
        failed_ = test_name in failed

        ran_val    = "YES" if ran    else "NO"
        passed_val = "YES" if passed_ else ("NO" if (ran and failed_) else "")

        cmd = (_BASE_CMD + cmd_suffix + '" -v -s') if cmd_suffix else ""

        row_fill = _PASS_FILL if passed_ else (_FAIL_FILL if failed_ else (_SKIP_FILL if (ran and not passed_ and not failed_) else _GREY_FILL))

        values = [idx, cls, test_name, what, ran_val, passed_val, cmd]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = _BORDER
            if ci in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ran:
                cell.fill = row_fill
            else:
                cell.fill = _GREY_FILL

        ws.row_dimensions[row_num].height = 30
        row_num += 1

    _col_width(ws, "A", 5)
    _col_width(ws, "B", 28)
    _col_width(ws, "C", 52)
    _col_width(ws, "D", 55)
    _col_width(ws, "E", 7)
    _col_width(ws, "F", 9)
    _col_width(ws, "G", 80)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # summary row
    total   = len(_REPORT_META)
    n_pass  = sum(1 for _, t, *_ in _REPORT_META if t in passed)
    n_fail  = sum(1 for _, t, *_ in _REPORT_META if t in failed)
    n_ran   = sum(1 for _, t, *_ in _REPORT_META if t in passed or t in failed or t in skipped)
    ws.append([])
    summary_row = row_num + 1
    ws.cell(summary_row, 1, "Summary").font = Font(bold=True)
    ws.cell(summary_row, 2, f"Ran: {n_ran}/{total}   Passed: {n_pass}   Failed: {n_fail}").font = Font(bold=True)
    ws.cell(summary_row, 3, f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    out_dir  = os.path.dirname(__file__)
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"direct_pb_report_{ts}.xlsx")
    wb.save(out_path)
    terminalreporter.write_sep("-", f"Excel report saved → {out_path}")


@pytest.fixture(scope="session")
def playwright_instance():
    with sync_playwright() as p:
        yield p


@pytest.fixture(scope="session")
def browser(playwright_instance):
    b = playwright_instance.chromium.launch(headless=False, slow_mo=150, args=["--start-maximized"])
    yield b
    b.close()


@pytest.fixture(scope="session")
def browser_context(browser):
    ctx = browser.new_context(no_viewport=True)
    yield ctx
    ctx.close()


@pytest.fixture(scope="session")
def logged_in_page(browser_context):
    """Login as kedar@rhythmflows.com and select tenant 'Eco Green Pvt Ltd'."""
    page = browser_context.new_page()
    page.goto(LOGIN_URL)

    page.wait_for_selector("input[name='Username']", timeout=15000)
    page.fill("input[name='Username']", EMAIL)
    page.fill("input[name='Password']", PASSWORD)
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)

    page.wait_for_selector("mat-select[aria-label], mat-select", timeout=15000)
    page.locator("mat-select").first.click(force=True)
    page.wait_for_selector(".dd-search-input", timeout=10000)
    page.locator(".dd-search-input").fill(TENANT)
    page.wait_for_timeout(800)
    for opt in page.locator(".mat-mdc-select-panel mat-option span.mdc-list-item__primary-text").all():
        if opt.inner_text().strip() == TENANT:
            opt.click(force=True)
            break
    # Wait for the dropdown overlay to close before clicking submit
    try:
        page.wait_for_selector(".mat-mdc-select-panel", state="hidden", timeout=3000)
    except Exception:
        page.keyboard.press("Escape")
        page.wait_for_timeout(300)
    page.wait_for_timeout(300)

    page.locator("button[type='submit']").click(force=True)
    page.wait_for_url(
        lambda url: "signin" not in url.lower() and "authentication" not in url.lower(),
        timeout=20000,
    )
    page.wait_for_timeout(1500)

    yield page
    page.close()


@pytest.fixture(scope="function")
def pb_page(logged_in_page):
    p = DirectPBPlaywrightPage(logged_in_page)
    p.navigate_to_page()
    yield p
    try:
        p.close_popup()
    except Exception:
        pass
