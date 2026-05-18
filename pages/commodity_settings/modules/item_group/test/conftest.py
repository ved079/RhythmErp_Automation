"""
conftest.py
-----------
Pytest fixtures for Item Group automation.

Fixtures:
  - driver          : session-scoped Edge browser
  - logged_in_driver: session-scoped, logs into RhythmERP
  - ig_page         : per-test, navigates to Item Group page

Also includes:
  - CSReportStore for Excel report generation
  - Log capture for report
  - Known issues tracking
  - pytest_sessionfinish hook for report output
"""

import os
import sys
import pytest
import logging
from datetime import datetime

PROJECT_ROOT = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "..")
)
sys.path.insert(0, PROJECT_ROOT)

from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService

from common.logger import log
from common.login_page import LoginPage
from config import RHYTHMERP_BASE_URL, RHYTHMERP_LOGIN_URL
from pages.commodity_settings.modules.item_group.item_group_page import ItemGroupPage


# ═══════════════════════════════════════════════════════════════
#  CSReportStore — lightweight Excel report generator
# ═══════════════════════════════════════════════════════════════

class CSReportStore:
    """Collects test results and generates an Excel report."""

    def __init__(self, screen_name="Item Group"):
        self.screen_name = screen_name
        self.results = []
        self.known_issues = {}

    def record(self, test_id, description, status, actual="", bug_ref=""):
        """Record a single test result."""
        self.results.append({
            "test_id": test_id,
            "description": description,
            "status": status,
            "actual": actual,
            "bug_ref": bug_ref,
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    def record_issue(self, test_id, description, actual, bug_ref=""):
        """Record a known issue (will be marked in report)."""
        self.known_issues[test_id] = {
            "description": description,
            "actual": actual,
            "bug_ref": bug_ref,
        }

    def generate_report(self, output_dir):
        """Generate an Excel report with all test results."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Item Group Test Results"

            # Styling
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="002F5496", end_color="002F5496", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="top")

            pass_fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
            fail_fill = PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid")
            skip_fill = PatternFill(start_color="00FFEB9C", end_color="00FFEB9C", fill_type="solid")
            xfail_fill = PatternFill(start_color="00F4B084", end_color="00F4B084", fill_type="solid")

            data_font = Font(name="Calibri", size=10)
            data_align = Alignment(vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Title
            ws.merge_cells("A1:I1")
            ws["A1"] = "Item Group — Automation Test Results"
            ws["A1"].font = Font(name="Calibri", size=14, bold=True)

            # Headers
            headers = ["#", "Test ID", "Screen Name", "Description", "Steps",
                       "Expected Result", "Actual Result", "Status", "Date"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Data rows
            for row_idx, result in enumerate(self.results, 4):
                status = result["status"]
                if status == "PASS":
                    fill = pass_fill
                elif status == "FAIL":
                    fill = fail_fill
                elif status == "SKIP":
                    fill = skip_fill
                elif status == "XFAIL":
                    fill = xfail_fill
                else:
                    fill = PatternFill()

                row_data = [
                    row_idx - 3,
                    result["test_id"],
                    self.screen_name,
                    result["description"],
                    "",
                    "",
                    result.get("actual", ""),
                    status,
                    result["date"],
                ]

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.fill = fill
                    cell.border = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 40
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 30
            ws.column_dimensions["G"].width = 30
            ws.column_dimensions["H"].width = 10
            ws.column_dimensions["I"].width = 20

            # Save
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"CommonSettings_Report_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            wb.save(filepath)
            print(f"\n{'=' * 60}")
            print(f"  ITEM GROUP REPORT GENERATED: {filepath}")
            print(f"{'=' * 60}\n")
            return filepath

        except Exception as e:
            log.warning(f"Report generation failed: {e}")
            return ""


# Global report store instance
report_store = CSReportStore("Item Group")


# ═══════════════════════════════════════════════════════════════
#  Known Issues for Item Group
# ═══════════════════════════════════════════════════════════════

report_store.record_issue(
    "IG-D01",
    "Duplicate Code ALLOWED — no uniqueness constraint",
    "ERP accepts duplicate Code values without error",
    "BUG-IG-001",
)
report_store.record_issue(
    "IG-D02",
    "Exact duplicate (same Code + Description) ALLOWED",
    "ERP accepts exact duplicate records without error",
    "BUG-IG-002",
)


# ═══════════════════════════════════════════════════════════════
#  Browser fixtures
# ═══════════════════════════════════════════════════════════════

@pytest.fixture(scope="session")
def driver():
    """Launch Edge browser (session-scoped)."""
    log.info("=" * 60)
    log.info("LAUNCHING BROWSER (RhythmERP - Item Group Tests)...")
    log.info("=" * 60)

    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--inprivate")

    # Suppress logging
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)

    log.info("Setting up Edge browser...")
    driver = webdriver.Edge(options=options)
    driver.implicitly_wait(5)
    log.info("Edge browser launched successfully (fresh profile)")

    yield driver

    log.info("=" * 60)
    log.info("CLOSING BROWSER...")
    log.info("=" * 60)
    try:
        driver.quit()
    except Exception:
        pass


@pytest.fixture(scope="session")
def logged_in_driver(driver):
    """Log into RhythmERP (session-scoped)."""
    log.info("=" * 60)
    log.info("LOGGING INTO RHYTHMERP...")
    log.info("=" * 60)

    login_page = LoginPage(driver)
    login_page.login()

    log.info("RhythmERP login successful!")
    return driver


@pytest.fixture
def ig_page(logged_in_driver):
    """Create ItemGroupPage instance and navigate to Item Group (per-test)."""
    page = ItemGroupPage(logged_in_driver)
    page.navigate_to_page()
    yield page


# ═══════════════════════════════════════════════════════════════
#  Log capture for report
# ═══════════════════════════════════════════════════════════════

class LogCapture(logging.Handler):
    """Custom log handler to capture log messages for report."""

    def __init__(self):
        super().__init__()
        self.records = []

    def emit(self, record):
        self.records.append(record)


_log_capture = LogCapture()
logging.getLogger("PACS_Automation").addHandler(_log_capture)


# ═══════════════════════════════════════════════════════════════
#  pytest hooks for report
# ═══════════════════════════════════════════════════════════════

@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    if report.when == "call":
        test_id = item.name.split("::")[-1].replace("test_", "").upper()
        description = item.funcargs.get("ig_page", None)
        desc_text = item.obj.__doc__ or ""

        if report.passed:
            report_store.record(test_id, desc_text.strip(), "PASS")
        elif report.failed:
            report_store.record(
                test_id, desc_text.strip(), "FAIL",
                actual=str(report.longrepr) if report.longrepr else ""
            )
        elif report.skipped:
            report_store.record(test_id, desc_text.strip(), "SKIP")


def pytest_sessionfinish(session, exitstatus):
    """Generate Excel report at session end."""
    reports_dir = os.path.join(
        os.path.dirname(__file__), "..", "reports"
    )
    report_store.generate_report(os.path.abspath(reports_dir))