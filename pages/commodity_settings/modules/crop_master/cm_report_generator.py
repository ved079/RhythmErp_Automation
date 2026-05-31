"""
cm_report_generator.py
-----------------------
Excel report generator for Crop Master test results.

Creates a formatted Excel file with:
  - Sheet 1: "Test Results" — one row per test case
  - Sheet 2: "Known Issues" — pre-registered known bugs
  - Sheet 3: "Summary" — pass/fail counts by category
"""

import os
import sys
from datetime import datetime

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
sys.path.insert(0, PROJECT_ROOT)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


class CMReportStore:
    """Store test results and known issues; generate Excel report."""

    def __init__(self):
        self.results = []
        self.known_issues = []
        self.start_time = None
        self.end_time = None

    def start(self):
        self.start_time = datetime.now()

    def finish(self):
        self.end_time = datetime.now()

    def add_result(self, test_id, test_name, category, status,
                   error="", duration=0.0, details=""):
        self.results.append({
            "test_id": test_id,
            "test_name": test_name,
            "category": category,
            "status": status,
            "error": error,
            "duration": round(duration, 2),
            "details": details,
        })

    def add_known_issue(self, issue_id, phase, description,
                        expected, actual, severity):
        self.known_issues.append({
            "issue_id": issue_id,
            "phase": phase,
            "description": description,
            "expected": expected,
            "actual": actual,
            "severity": severity,
        })

    def generate_excel(self, output_dir=None):
        """Generate the Excel report and return the file path."""
        if output_dir is None:
            output_dir = os.path.join(PROJECT_ROOT, "reports")
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"CropMaster_Report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()

        # ---- Styles ----
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        pass_font = Font(name="Calibri", size=11, color="006100")
        fail_font = Font(name="Calibri", size=11, color="9C0006")
        warn_font = Font(name="Calibri", size=11, color="9C6500")
        normal_font = Font(name="Calibri", size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ====== Sheet 1: Test Results ======
        ws = wb.active
        ws.title = "Test Results"

        ws.merge_cells("A1:G1")
        ws["A1"].value = "Crop Master — Automation Test Results"
        ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="2F5496")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        start_str = self.start_time.strftime("%Y-%m-%d %H:%M:%S") if self.start_time else "N/A"
        end_str = self.end_time.strftime("%Y-%m-%d %H:%M:%S") if self.end_time else "N/A"
        ws["A2"].value = (
            f"Start: {start_str} | End: {end_str} | "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        ws["A2"].font = Font(name="Calibri", size=10, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Test ID", "Test Name", "Category", "Status", "Error", "Duration (s)", "Details"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, result in enumerate(self.results, 5):
            values = [
                result.get("test_id", ""),
                result.get("test_name", ""),
                result.get("category", ""),
                result.get("status", ""),
                result.get("error", ""),
                result.get("duration", 0),
                result.get("details", ""),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [5, 7]))
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if value == "PASSED":
                        cell.fill = pass_fill
                        cell.font = pass_font
                    elif value == "FAILED":
                        cell.fill = fail_fill
                        cell.font = fail_font
                    elif value == "XFAIL":
                        cell.fill = warn_fill
                        cell.font = warn_font

        col_widths = [12, 45, 22, 12, 50, 14, 50]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # ====== Sheet 2: Known Issues ======
        ws2 = wb.create_sheet(title="Known Issues")

        ws2.merge_cells("A1:F1")
        ws2["A1"].value = "Crop Master — Known Issues"
        ws2["A1"].font = Font(name="Calibri", bold=True, size=16, color="2F5496")
        ws2["A1"].alignment = Alignment(horizontal="center")

        ki_headers = ["Issue ID", "Phase", "Description", "Expected", "Actual", "Severity"]
        for col_idx, header in enumerate(ki_headers, 1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        severity_fills = {
            "Critical": PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid"),
            "High": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            "Medium": warn_fill,
            "Low-Medium": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
            "Low": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        }
        severity_fonts = {
            "Critical": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "High": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "Medium": warn_font,
            "Low-Medium": Font(name="Calibri", size=11, color="9C6500"),
            "Low": Font(name="Calibri", size=11, color="2F5496"),
        }

        for row_idx, issue in enumerate(self.known_issues, 4):
            values = [
                issue.get("issue_id", ""),
                issue.get("phase", ""),
                issue.get("description", ""),
                issue.get("expected", ""),
                issue.get("actual", ""),
                issue.get("severity", ""),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [3, 4, 5]))
                if col_idx == 6 and value in severity_fills:
                    cell.fill = severity_fills[value]
                    cell.font = severity_fonts[value]
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ki_col_widths = [12, 14, 40, 40, 45, 14]
        for idx, width in enumerate(ki_col_widths, 1):
            ws2.column_dimensions[get_column_letter(idx)].width = width

        # ====== Sheet 3: Summary ======
        ws3 = wb.create_sheet(title="Summary")

        ws3.merge_cells("A1:D1")
        ws3["A1"].value = "Test Summary by Category"
        ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

        categories = {}
        for r in self.results:
            cat = r.get("category", "Unknown")
            if cat not in categories:
                categories[cat] = {"passed": 0, "failed": 0, "total": 0}
            categories[cat]["total"] += 1
            if r.get("status") == "PASSED":
                categories[cat]["passed"] += 1
            else:
                categories[cat]["failed"] += 1

        summary_headers = ["Category", "Total", "Passed", "Failed"]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws3.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        row_idx = 4
        total_all = passed_all = failed_all = 0
        for cat, counts in categories.items():
            ws3.cell(row=row_idx, column=1, value=cat).font = normal_font
            ws3.cell(row=row_idx, column=1).border = thin_border
            ws3.cell(row=row_idx, column=2, value=counts["total"]).font = normal_font
            ws3.cell(row=row_idx, column=2).border = thin_border
            ws3.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
            ws3.cell(row=row_idx, column=3, value=counts["passed"]).font = pass_font
            ws3.cell(row=row_idx, column=3).fill = pass_fill
            ws3.cell(row=row_idx, column=3).border = thin_border
            ws3.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
            ws3.cell(row=row_idx, column=4, value=counts["failed"]).font = fail_font if counts["failed"] > 0 else normal_font
            ws3.cell(row=row_idx, column=4).fill = fail_fill if counts["failed"] > 0 else PatternFill()
            ws3.cell(row=row_idx, column=4).border = thin_border
            ws3.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
            total_all += counts["total"]
            passed_all += counts["passed"]
            failed_all += counts["failed"]
            row_idx += 1

        row_idx += 1
        ws3.cell(row=row_idx, column=1, value="GRAND TOTAL").font = Font(name="Calibri", bold=True, size=12)
        ws3.cell(row=row_idx, column=1).border = thin_border
        ws3.cell(row=row_idx, column=2, value=total_all).font = Font(name="Calibri", bold=True, size=12)
        ws3.cell(row=row_idx, column=2).border = thin_border
        ws3.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws3.cell(row=row_idx, column=3, value=passed_all).font = Font(name="Calibri", bold=True, size=12, color="006100")
        ws3.cell(row=row_idx, column=3).fill = pass_fill
        ws3.cell(row=row_idx, column=3).border = thin_border
        ws3.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
        ws3.cell(row=row_idx, column=4, value=failed_all).font = Font(name="Calibri", bold=True, size=12, color="9C0006")
        ws3.cell(row=row_idx, column=4).fill = fail_fill if failed_all > 0 else PatternFill()
        ws3.cell(row=row_idx, column=4).border = thin_border
        ws3.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")

        row_idx += 1
        pass_rate = (passed_all / total_all * 100) if total_all > 0 else 0
        ws3.merge_cells(f"A{row_idx}:D{row_idx}")
        ws3.cell(row=row_idx, column=1, value=f"Pass Rate: {pass_rate:.1f}%").font = Font(
            name="Calibri", bold=True, size=14, color="2F5496",
        )

        ws3.column_dimensions["A"].width = 24
        ws3.column_dimensions["B"].width = 12
        ws3.column_dimensions["C"].width = 12
        ws3.column_dimensions["D"].width = 12

        wb.save(filepath)
        print(f"\n  ============================================================\n  REPORT GENERATED: {filepath}\n  ============================================================")
        return filepath


# Global report store instance
cm_report = CMReportStore()


def generate_cm_report(output_dir=None):
    """Generate the Excel report from the global report store."""
    cm_report.finish()
    return cm_report.generate_excel(output_dir)