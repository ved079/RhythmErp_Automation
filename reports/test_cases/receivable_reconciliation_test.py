import time
import os
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import logging
import hashlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up module-level logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.handlers:
    console = logging.StreamHandler()
    console.setFormatter(logging.Formatter('%(asctime)s | %(levelname)-8s | %(message)s', datefmt='%H:%M:%S'))
    logger.addHandler(console)

# ==========================================
# 1. UI INTERACTION HELPERS (same as payable)
# ==========================================

def select_dropdown(driver, wait, value, control_name=None, label_text=None, control_id=None, searchable=True):
    """Universal dropdown selection."""
    try:
        identifier = control_name or label_text or control_id or "unknown"
        logger.info(f"➡️ Selecting {identifier}: {value}")

        dropdown = None
        if control_name:
            dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"mat-select[formcontrolname='{control_name}']")))
        elif label_text:
            dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, f"//mat-label[contains(text(), '{label_text}')]/ancestor::mat-form-field//mat-select")))
        elif control_id:
            dropdown = wait.until(EC.element_to_be_clickable((By.ID, control_id)))

        if not dropdown:
            raise ValueError("No locator provided")

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown)
        driver.execute_script("arguments[0].click();", dropdown)

        overlay = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "body > .cdk-overlay-container .cdk-overlay-pane:last-child")
        ))
        wait.until(EC.visibility_of(overlay))

        if searchable:
            search_input = wait.until(EC.presence_of_element_located(
                (By.XPATH, ".//input[@placeholder='Search' or contains(@class,'mat-filter-input')]")
            ))
            search_input.clear()
            search_input.send_keys(value)
            time.sleep(1)

        try:
            option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//mat-option//span[normalize-space()='{value}']")))
        except:
            option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//mat-option//span[contains(text(), '{value}')]")))

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", option)
        driver.execute_script("arguments[0].click();", option)
        
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-backdrop")))
        time.sleep(0.5)
    except Exception as e:
        logger.error(f"❌ Dropdown failed for {identifier}: {e}")
        driver.save_screenshot(f"dropdown_error_{identifier}.png")
        raise

def go_to_reports_page(driver, wait):
    logger.info("Navigating to Reports page...")
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-container, .swal2-container, .ngx-spinner-overlay")))
    except:
        pass 

    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)

    xpath = "//a[contains(@href, 'rhythm-report/reports')] | //span[contains(text(), 'All Reports')]/ancestor::a"
    reports_menu = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", reports_menu)
    driver.execute_script("arguments[0].click();", reports_menu) 

    time.sleep(2)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "mat-select[formcontrolname='report_name']")))

def select_report_name(driver, wait, report_name="Receivable"):
    logger.info(f"   🔽 Selecting Report Name: {report_name}...")
    dropdown = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "mat-select[formcontrolname='report_name']")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown)
    time.sleep(0.5)
    
    if report_name in dropdown.text:
        return True

    driver.execute_script("arguments[0].click();", dropdown)
    overlay = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "body > .cdk-overlay-container .cdk-overlay-pane:last-child")
    ))
    wait.until(EC.visibility_of(overlay))

    option = wait.until(EC.presence_of_element_located((By.XPATH, f"//mat-option//span[contains(normalize-space(), '{report_name}')]")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", option)
    driver.execute_script("arguments[0].click();", option)
    
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-backdrop")))
    time.sleep(1)
    return True

def fill_receivable_form(driver, wait, data):
    logger.info("📝 Filling Receivable form...")
    select_report_name(driver, wait, "Receivable")
    time.sleep(1.5) 
    wait.until(EC.presence_of_element_located((By.ID, "file_format")))
    
    if data.get('file_format'):
        try:
            select_dropdown(driver, wait, value=data['file_format'], control_id='file_format', searchable=False)
        except Exception:
            select_dropdown(driver, wait, value=data['file_format'], control_name='file_format', searchable=False)

def click_view(driver, wait):
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "mat-spinner, .ngx-spinner-overlay, .cdk-overlay-backdrop")))
    except:
        pass

    view_btn = wait.until(EC.presence_of_element_located((By.XPATH, "//button[contains(.,'View')]")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", view_btn)
    driver.execute_script("arguments[0].click();", view_btn)
    logger.info("   ✅ View button clicked")
    
    logger.info("⏳ Waiting for report table to load...")
    time.sleep(3) 
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table, .report-container")))
    logger.info("✅ Report table loaded.")

def click_download(driver, wait):
    try:
        download_xpath = "//button[contains(normalize-space(), 'Download')] | //button[contains(@class, 'apply') and contains(text(), 'Download')]"
        download_btn = wait.until(EC.presence_of_element_located((By.XPATH, download_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", download_btn)
        driver.execute_script("arguments[0].click();", download_btn)
        time.sleep(4)
        logger.info("✅ Download triggered successfully.")
    except Exception as e:
        logger.warning(f"⚠️ Download failed: {e}")
        driver.save_screenshot("download_button_error.png")

# ==========================================
# 2. RECONCILIATION & MATH HELPERS
# ==========================================

def clean_amount(amount_str):
    if not amount_str or amount_str.strip() == "": return 0.0
    try: return float(amount_str.replace(',', '').strip())
    except ValueError: return 0.0

def set_pagination_max(driver, wait):
    logger.info("🚀 Forcing pagination to 1000 items for deep-dive audit...")
    try:
        dropdown = wait.until(EC.presence_of_element_located((By.ID, "itemsPerPage")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown)
        driver.execute_script("arguments[0].value = '1000';", dropdown)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", dropdown)
        try:
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "mat-spinner, .ngx-spinner-overlay")))
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "mat-spinner, .ngx-spinner-overlay")))
        except: time.sleep(2)
        logger.info("✅ Pagination set. All rows rendered.")
    except Exception as e:
        logger.warning(f"⚠️ Pagination adjustment failed, scraping current page: {e}")

def extract_nested_ledger(driver, parent_row):
    """Extract child ledger data from expanded row (Receivable version)."""
    nested_table = parent_row.find_element(By.XPATH, "following-sibling::tr[1]//table")
    rows = nested_table.find_elements(By.CSS_SELECTOR, "tbody > tr")
    
    ledger_data = {
        "opening_cr": 0.0, "opening_dr": 0.0,
        "transaction_cr_sum": 0.0, "transaction_dr_sum": 0.0,
        "closing_cr": 0.0, "closing_dr": 0.0
    }
    
    for row in rows:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) < 5: continue
            
        desc = cols[2].text.strip()   # Ref No / Description column
        cr = clean_amount(cols[3].text)   # Credit column
        dr = clean_amount(cols[4].text)   # Debit column
        
        if "Opening Balance" in desc:
            ledger_data["opening_cr"] = cr
            ledger_data["opening_dr"] = dr
        elif "Closing Balance" in desc:
            ledger_data["closing_cr"] = cr
            ledger_data["closing_dr"] = dr
        else:
            ledger_data["transaction_cr_sum"] += cr
            ledger_data["transaction_dr_sum"] += dr
            
    return ledger_data

# ==========================================
# 3. MAIN EXECUTOR 
# ==========================================

def run_receivable_reconciliation(driver, wait, data):
    """Enhanced reconciliation for Receivable report."""
    logger.info("\n--- INITIATING RECEIVABLE RECONCILIATION SUITE ---")
    start_time = time.time()
    
    # Audit log
    audit_log = []
    def log_action(action):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_log.append({"Timestamp": timestamp, "Action": action})
        logger.info(f"   [LOG] {timestamp} - {action}")
    
    log_action("Start reconciliation suite")
    
    # UI Navigation
    log_action("Navigate to Reports page")
    go_to_reports_page(driver, wait)
    log_action("Fill Receivable form")
    fill_receivable_form(driver, wait, data)
    log_action("Click View button")
    click_view(driver, wait)
    
    # Setup
    log_action("Set pagination to 1000")
    set_pagination_max(driver, wait)
    test_results = []
    raw_child_data = []
    
    # Find parent rows (those with expand icon)
    parent_rows = driver.find_elements(By.XPATH, "//tr[.//i[contains(@class, 'fa-angle-right') or contains(@class, 'fa-angle-down')]]")
    num_rows = len(parent_rows)
    logger.info(f"📊 Found {num_rows} customers to audit. Commencing row checks...")
    log_action(f"Found {num_rows} parent rows")
    
    # Loop over each customer row
    for index in range(num_rows):
        try:
            # Re-fetch rows each iteration to avoid stale references
            current_rows = driver.find_elements(By.XPATH, "//tr[.//i[contains(@class, 'fa-angle-right') or contains(@class, 'fa-angle-down')]]")
            row = current_rows[index]
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) < 8:
                log_action(f"Row {index+1} skipped – insufficient columns")
                continue
                
            customer_name = cols[1].text.strip()
            p_open_cr = clean_amount(cols[2].text)
            p_open_dr = clean_amount(cols[3].text)
            p_txn_cr = clean_amount(cols[4].text)
            p_txn_dr = clean_amount(cols[5].text)
            p_close_cr = clean_amount(cols[6].text)
            p_close_dr = clean_amount(cols[7].text)
            
            log_action(f"Row {index+1}: Expanding {customer_name}")
            toggle_icon = row.find_element(By.CSS_SELECTOR, "i.fa-angle-right, i.fa-angle-down")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", toggle_icon)
            driver.execute_script("arguments[0].click();", toggle_icon)
            time.sleep(1)
            
            # Extract child data and also capture raw rows
            nested_table = row.find_element(By.XPATH, "following-sibling::tr[1]//table")
            child_rows = nested_table.find_elements(By.CSS_SELECTOR, "tbody > tr")
            
            child_data = {
                "opening_cr": 0.0, "opening_dr": 0.0,
                "transaction_cr_sum": 0.0, "transaction_dr_sum": 0.0,
                "closing_cr": 0.0, "closing_dr": 0.0
            }
            
            # --- ENHANCED RAW CHILD DATA COLLECTION (with Running Balance & Counter) ---
            running_balance = 0.0
            txn_counter = 0
            
            for child_row in child_rows:
                child_cols = child_row.find_elements(By.TAG_NAME, "td")
                if len(child_cols) < 5: continue
                
                trans_date = child_cols[1].text.strip() if len(child_cols) > 1 else ""
                desc = child_cols[2].text.strip()
                cr = clean_amount(child_cols[3].text)
                dr = clean_amount(child_cols[4].text)
                
                # Determine row type
                is_opening = "Opening Balance" in desc
                is_closing = "Closing Balance" in desc
                is_transaction = not (is_opening or is_closing)
                
                # Update running balance only for opening and transaction rows
                if is_opening or is_transaction:
                    running_balance = running_balance + cr - dr
                
                # Increment transaction counter only for real transactions
                if is_transaction:
                    txn_counter += 1
                
                # Store row (always include all rows, even closing)
                raw_child_data.append({
                    "Customer": customer_name,
                    "Transaction Date": trans_date,
                    "Description": desc,
                    "Credit": cr,
                    "Debit": dr,
                    "Running Balance": round(running_balance, 2) if not is_closing else "",  # blank for closing row
                    "Transaction #": txn_counter if is_transaction else ""
                })
                
                # Math validation accumulation (unchanged)
                if is_opening:
                    child_data["opening_cr"] = cr
                    child_data["opening_dr"] = dr
                elif is_closing:
                    child_data["closing_cr"] = cr
                    child_data["closing_dr"] = dr
                else:
                    child_data["transaction_cr_sum"] += cr
                    child_data["transaction_dr_sum"] += dr
            
            log_action(f"Row {index+1}: Closing {customer_name}")
            # Re-locate the row to close it
            refreshed_rows = driver.find_elements(By.XPATH, "//tr[.//i[contains(@class, 'fa-angle-right') or contains(@class, 'fa-angle-down')]]")
            active_row = refreshed_rows[index]
            close_icon = active_row.find_element(By.CSS_SELECTOR, "i.fa-angle-down, i.fa-angle-right")
            driver.execute_script("arguments[0].click();", close_icon)
            time.sleep(0.5)
            
            # Assertions (unchanged)
            errors = []
            match_txn_cr = (round(child_data["transaction_cr_sum"], 2) == round(p_txn_cr, 2))
            match_txn_dr = (round(child_data["transaction_dr_sum"], 2) == round(p_txn_dr, 2))
            if not match_txn_cr:
                errors.append(f"Cr Mismatch: Parent {p_txn_cr} != Child {child_data['transaction_cr_sum']}")
            if not match_txn_dr:
                errors.append(f"Dr Mismatch: Parent {p_txn_dr} != Child {child_data['transaction_dr_sum']}")
            
            # Net balance check (same formula as Payable)
            expected_net = (p_open_cr - p_open_dr) + (p_txn_cr - p_txn_dr)
            if expected_net > 0:
                calc_close_cr, calc_close_dr = expected_net, 0.0
            elif expected_net < 0:
                calc_close_cr, calc_close_dr = 0.0, abs(expected_net)
            else:
                calc_close_cr, calc_close_dr = 0.0, 0.0
                
            math_integrity = (round(calc_close_cr,2) == round(p_close_cr,2)) and (round(calc_close_dr,2) == round(p_close_dr,2))
            if not math_integrity:
                errors.append(f"Math failed: Expected closing Cr/Dr {calc_close_cr}/{calc_close_dr} != UI {p_close_cr}/{p_close_dr}")
            
            row_status = "PASS" if (match_txn_cr and match_txn_dr and math_integrity) else "FAIL"
            
            test_results.append({
                "Customer Name": customer_name,
                "Parent_Txn_Cr": p_txn_cr, "Child_Sum_Cr": child_data["transaction_cr_sum"], "Match_Cr": match_txn_cr,
                "Parent_Txn_Dr": p_txn_dr, "Child_Sum_Dr": child_data["transaction_dr_sum"], "Match_Dr": match_txn_dr,
                "Math_Integrity": math_integrity, "Status": row_status, "Errors": " | ".join(errors) if errors else "None"
            })
            log_action(f"Row {index+1} {customer_name} → {row_status}")
            
        except Exception as e:
            logger.error(f"❌ Error auditing row {index + 1}: {e}")
            log_action(f"Row {index+1} ERROR: {str(e)[:100]}")
    
    # Global totals validation (unchanged)
    logger.info("\n🌐 Reconciling Global Headers...")
    log_action("Global header reconciliation")
    
    try:
        header_rows = driver.find_elements(By.CSS_SELECTOR, "tr.table-totals")
        # Assuming 3 rows: Total Opening Amount, Total Transaction Amount, Total Closing Amount
        ui_total_txn = clean_amount(header_rows[1].find_elements(By.TAG_NAME, "td")[-1].text)
        
        valid_rows = test_results
        our_total_txn_cr = sum(row['Parent_Txn_Cr'] for row in valid_rows)
        our_total_txn_dr = sum(row['Parent_Txn_Dr'] for row in valid_rows)
        our_net_txn = abs(our_total_txn_cr - our_total_txn_dr)
        global_match = (round(ui_total_txn, 2) == round(our_net_txn, 2))
        
        global_row = {
            "Customer Name": "GRAND TOTAL RECONCILIATION",
            "Parent_Txn_Cr": f"UI Total: {ui_total_txn}",
            "Child_Sum_Cr": f"Our Net Calc: {our_net_txn}",
            "Status": "PASS" if global_match else "FAIL",
            "Errors": "Global UI Header matches row sums!" if global_match else "CRITICAL: Global UI Header does NOT match row sums!"
        }
        test_results.append(global_row)
        log_action(f"Global header check: {global_row['Status']}")
    except Exception as e:
        logger.warning(f"⚠️ Could not extract global headers: {e}")
        log_action(f"Global header extraction failed: {e}")
        global_row = {"Customer Name": "ERROR", "Errors": str(e)}
    
    # Screenshot (unchanged)
    screenshot_dir = r"C:\Users\vedantd\Desktop\selenium files\reports\test_cases\screenshots"
    os.makedirs(screenshot_dir, exist_ok=True)
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    screenshot_path = os.path.join(screenshot_dir, f"receivable_report_{timestamp_str}.png")
    driver.save_screenshot(screenshot_path)
    logger.info(f"📸 Screenshot saved: {screenshot_path}")
    log_action(f"Screenshot captured: {screenshot_path}")
    
    # Checksum (unchanged)
    checksum_str = ""
    for row in test_results:
        for key in ["Customer Name", "Parent_Txn_Cr", "Child_Sum_Cr", "Match_Cr", "Parent_Txn_Dr", "Child_Sum_Dr", "Match_Dr", "Math_Integrity", "Status", "Errors"]:
            checksum_str += str(row.get(key, ""))
    checksum = hashlib.md5(checksum_str.encode('utf-8')).hexdigest()
    log_action(f"Data checksum: {checksum}")
    
    # Metadata (unchanged)
    metadata = [
        {"Attribute": "Report Timestamp", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Attribute": "Screenshot Path", "Value": screenshot_path},
        {"Attribute": "Data Checksum (MD5)", "Value": checksum},
        {"Attribute": "Total Customers Audited", "Value": num_rows},
        {"Attribute": "Test Duration (seconds)", "Value": f"{time.time() - start_time:.2f}"}
    ]
    
    # Multi-sheet Excel with styling (unchanged)
    logger.info("\n💾 Generating professional Audit Report Excel...")
    download_dir = r"C:\Users\vedantd\Desktop\selenium files\reports\test_cases\downloads"
    os.makedirs(download_dir, exist_ok=True)
    file_name = f"Receivable_Audit_Reconciliation_{timestamp_str}.xlsx"
    file_path = os.path.join(download_dir, file_name)
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        pd.DataFrame(test_results).to_excel(writer, sheet_name='Reconciliation_Summary', index=False)
        if raw_child_data:
            pd.DataFrame(raw_child_data).to_excel(writer, sheet_name='Raw_Child_Data', index=False)
        if audit_log:
            pd.DataFrame(audit_log).to_excel(writer, sheet_name='Audit_Trail', index=False)
        pd.DataFrame(metadata).to_excel(writer, sheet_name='Metadata', index=False)
        
        # Styling with Indian currency format
        indian_format = '#,##,##0.00'
        
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            
            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply Indian number format to amount columns
            if sheet_name == 'Reconciliation_Summary':
                # Columns that contain monetary values
                amount_headers = ['Parent_Txn_Cr', 'Child_Sum_Cr', 'Parent_Txn_Dr', 'Child_Sum_Dr']
                header_row = [cell.value for cell in ws[1]]
                for col_idx, header in enumerate(header_row, start=1):
                    if header in amount_headers:
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = indian_format
            
            elif sheet_name == 'Raw_Child_Data' and ws.max_row > 1:
                amount_headers = ['Credit', 'Debit', 'Running Balance']
                header_row = [cell.value for cell in ws[1]]
                for col_idx, header in enumerate(header_row, start=1):
                    if header in amount_headers:
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = indian_format
            
            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
            
            # Borders
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
    
    logger.info(f"✅ Full reconciliation complete! Audit saved to: {file_path}")
    log_action(f"Excel report generated: {file_path}")
    
    # Trigger system download (unchanged)
    click_download(driver, wait)
    log_action("System download triggered")